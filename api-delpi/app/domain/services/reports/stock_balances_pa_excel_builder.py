"""Gera XLSX de saldos PA (openpyxl) — espelha o export do Portal PCP."""

from __future__ import annotations

import base64
import io
from typing import Any, Mapping, Sequence

from openpyxl import Workbook
from openpyxl.styles import Border, Font, Side
from openpyxl.utils import get_column_letter

from app.domain.services.reports.report_types import ReportAttachment
from app.domain.services.reports.stock_balances_pa_rules import (
    COLUMN_PRODUCT,
    COLUMN_QUANTITY,
    XLSX_CONTENT_TYPE,
    export_file_base,
    export_quantity,
)

_THIN = Border(
    left=Side(style="thin", color="000000"),
    right=Side(style="thin", color="000000"),
    top=Side(style="thin", color="000000"),
    bottom=Side(style="thin", color="000000"),
)


class StockBalancesPaExcelBuilder:
    """Planilha com Produto + Quantidade×1000 e bordas em todas as células."""

    @staticmethod
    def build_bytes(
        rows: Sequence[Mapping[str, Any]],
        *,
        branch: str,
        sheet_title: str,
    ) -> bytes:
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = (sheet_title or "Saldos")[:31]

        headers = ("Produto", "Quantidade")
        sheet.append(list(headers))
        for cell in sheet[1]:
            cell.border = _THIN
            cell.font = Font(bold=True)

        for row in rows:
            product = str(row.get(COLUMN_PRODUCT) or "").strip()
            quantity = export_quantity(row.get(COLUMN_QUANTITY))
            sheet.append([product, quantity])
            for cell in sheet[sheet.max_row]:
                cell.border = _THIN

        for index, header in enumerate(headers, start=1):
            max_len = len(header)
            for row_cells in sheet.iter_rows(min_row=2, min_col=index, max_col=index):
                for cell in row_cells:
                    max_len = max(max_len, len(str(cell.value or "")))
            sheet.column_dimensions[get_column_letter(index)].width = min(max_len + 2, 50)

        buffer = io.BytesIO()
        workbook.save(buffer)
        return buffer.getvalue()

    @classmethod
    def build_attachment(
        cls,
        rows: Sequence[Mapping[str, Any]],
        *,
        branch: str,
        issued_on: Any = None,
    ) -> ReportAttachment:
        from app.domain.services.reports.stock_balances_pa_rules import title_for_branch

        title = title_for_branch(branch)
        payload = cls.build_bytes(rows, branch=branch, sheet_title=title)
        file_name = f"{export_file_base(branch, issued_on)}.xlsx"
        return ReportAttachment(
            name=file_name,
            content_type=XLSX_CONTENT_TYPE,
            content_base64=base64.b64encode(payload).decode("ascii"),
        )
