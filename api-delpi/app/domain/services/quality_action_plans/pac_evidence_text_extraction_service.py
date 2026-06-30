from __future__ import annotations

import io
from typing import Any

from openpyxl import load_workbook

_MAX_EXCERPT_CHARS = 80_000


def _decode_text(content: bytes) -> str:
    for encoding in ("utf-8", "latin-1", "cp1252"):
        try:
            return content.decode(encoding)
        except UnicodeDecodeError:
            continue
    return content.decode("utf-8", errors="replace")


def _extract_docx_text(content: bytes) -> str:
    import xml.etree.ElementTree as ET
    import zipfile

    with zipfile.ZipFile(io.BytesIO(content)) as archive:
        xml_bytes = archive.read("word/document.xml")
    root = ET.fromstring(xml_bytes)
    namespace = "{http://schemas.openxmlformats.org/wordprocessingml/2006/main}"
    paragraphs: list[str] = []
    for paragraph in root.iter(f"{namespace}p"):
        parts = [node.text for node in paragraph.iter(f"{namespace}t") if node.text]
        line = "".join(parts).strip()
        if line:
            paragraphs.append(line)
    return "\n\n".join(paragraphs)


def _extract_xlsx_text(content: bytes) -> str:
    workbook = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    parts: list[str] = []
    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        parts.append(f"## Planilha: {sheet_name}")
        row_count = 0
        for row in sheet.iter_rows(values_only=True):
            cells = [str(cell).strip() for cell in row if cell is not None and str(cell).strip()]
            if not cells:
                continue
            parts.append(" | ".join(cells))
            row_count += 1
            if row_count >= 500:
                parts.append("… (planilha truncada após 500 linhas)")
                break
    workbook.close()
    return "\n".join(parts)


def extract_evidence_text(
    *,
    content: bytes,
    mime_type: str | None,
    file_name: str | None,
) -> dict[str, Any]:
    """Extrai texto legível de evidências PAC para agentes e busca."""
    mime = (mime_type or "").lower().strip()
    name = (file_name or "").lower().strip()
    extracted = ""
    format_label = "binary"

    if mime.startswith("text/") or name.endswith((".txt", ".csv")):
        extracted = _decode_text(content)
        format_label = "text"
    elif mime in {
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        "application/vnd.ms-excel",
    } or name.endswith((".xlsx", ".xls")):
        extracted = _extract_xlsx_text(content)
        format_label = "spreadsheet"
    elif mime in {
        "application/vnd.openxmlformats-officedocument.wordprocessingml.document",
        "application/msword",
    } or name.endswith((".docx", ".doc")):
        if name.endswith(".docx") or "wordprocessingml" in mime:
            extracted = _extract_docx_text(content)
            format_label = "document"
        else:
            format_label = "document"
            extracted = ""
    elif mime == "application/pdf" or name.endswith(".pdf"):
        format_label = "pdf"
        extracted = ""

    normalized = extracted.strip()
    truncated = len(normalized) > _MAX_EXCERPT_CHARS
    if truncated:
        normalized = normalized[:_MAX_EXCERPT_CHARS] + "\n… (conteúdo truncado)"

    return {
        "format": format_label,
        "text_content": normalized,
        "char_count": len(normalized),
        "truncated": truncated,
        "extractable": bool(normalized) or format_label in {"text", "spreadsheet", "document"},
    }
