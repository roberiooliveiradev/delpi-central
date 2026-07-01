from __future__ import annotations

import io
from datetime import date, datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.drawing.image import Image as XlImage

from app.domain.services.quality_action_plans.action_responsibles_service import (
    format_responsible_display_name,
    responsibles_from_legacy_action,
)
from app.domain.services.quality_action_plans.quality_action_plan_contact_roles_service import (
    resolve_customer_contact_email,
    resolve_customer_contact_name,
    resolve_delpi_contact_area,
    resolve_delpi_contact_phone,
    resolve_delpi_primary_contact_name,
    resolve_delpi_quality_contact,
    format_delpi_contact_area_label,
)
from app.domain.services.quality_action_plans.five_whys_service import format_why_step_answer_cell
from app.domain.services.quality_action_plans.rnc_8d_quantity_field_service import (
    resolve_quantity_display,
)
from app.domain.services.quality_action_plans.rnc_8d_export_template_service import (
    resolve_export_template_key_for_plan,
    resolve_export_template_path,
)
from app.domain.services.quality_action_plans.rnc_8d_template_fill_service import (
    fill_template_workbook,
    load_merged_cell_anchors,
    overlay_zip_parts,
    put_cell_value,
)

TEMPLATE_FIXTURE_PATH = (
    Path(__file__).resolve().parents[4]
    / "tests"
    / "fixtures"
    / "quality"
    / "rnc_8d_template_minimal.xlsx"
)

SUPPLIER_BY_BRANCH = {
    "01": "12243 - Delpi Componentes Ltda EPP",
    "02": "12243 - Delpi Componentes Ltda EPP",
}

ANNEX_SHEET_CANDIDATES = ("Anexos(Evidencias)", "Anexos", "Attachment")
ANNEX_IMAGE_MAX_WIDTH_PX = 480
IMAGE_MIME_PREFIX = "image/"
WEG_TEMPLATE_KEY = "weg_wfr20997"


def resolve_rnc_8d_template_path(template_key: str | None = None) -> Path:
    key = template_key or WEG_TEMPLATE_KEY
    try:
        return resolve_export_template_path(key)
    except (KeyError, FileNotFoundError):
        if TEMPLATE_FIXTURE_PATH.is_file():
            return TEMPLATE_FIXTURE_PATH
        raise


def _excel_date(value: str | date | datetime | None) -> date | None:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    text = str(value).strip()
    if not text:
        return None
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y"):
        try:
            return datetime.strptime(text[:10], fmt).date()
        except ValueError:
            continue
    try:
        return datetime.fromisoformat(text.replace("Z", "+00:00")).date()
    except ValueError:
        return None


def _format_date_value(value: str | date | datetime | None) -> str | None:
    parsed = _excel_date(value)
    if parsed is None:
        return None
    return parsed.strftime("%d/%m/%Y")


def _action_responsible_label(action: dict[str, Any]) -> str | None:
    label = format_responsible_display_name(responsibles_from_legacy_action(action))
    return label or None


def is_image_evidence(evidence: dict[str, Any]) -> bool:
    mime = str(evidence.get("mime_type") or "").lower()
    if mime.startswith(IMAGE_MIME_PREFIX):
        return True
    evidence_type = str(evidence.get("type") or "").lower()
    return evidence_type == "image"


def _resolve_annex_sheet(wb) -> Any | None:
    for name in ANNEX_SHEET_CANDIDATES:
        if name in wb.sheetnames:
            return wb[name]
    return None


def _embed_annex_images(ws: Any, image_annexes: list[dict[str, Any]]) -> None:
    row = 3
    for item in image_annexes:
        content = item.get("content")
        if not content:
            continue
        try:
            img = XlImage(io.BytesIO(content))
        except Exception:
            continue
        if img.width > ANNEX_IMAGE_MAX_WIDTH_PX:
            ratio = ANNEX_IMAGE_MAX_WIDTH_PX / img.width
            img.width = int(img.width * ratio)
            img.height = int(img.height * ratio)
        ws.add_image(img, f"A{row}")
        label = str(item.get("file_name") or "").strip()
        description = str(item.get("description") or "").strip()
        if label:
            ws[f"D{row}"] = label
        if description:
            ws[f"D{row + 1}"] = description
        row += max(12, int(img.height / 15) + 2)


def _material_label(plan: dict[str, Any]) -> str:
    code = (plan.get("product_code") or "").strip()
    desc = (plan.get("product_description") or "").strip()
    if code and desc:
        return f"{code} - {desc}"
    return code or desc


def collect_image_annexes_for_export(
    *,
    plan_id: str,
    evidences: list[dict[str, Any]],
    storage: Any,
) -> list[dict[str, Any]]:
    annexes: list[dict[str, Any]] = []
    for evidence in evidences:
        if not is_image_evidence(evidence):
            continue
        stored_name = str(evidence.get("stored_name") or "").strip()
        if not stored_name:
            continue
        try:
            file_path = storage.resolve_evidence_file(
                stored_name=stored_name,
                plan_id_candidates=storage.plan_id_candidates(
                    plan_ref=plan_id,
                    evidence=evidence,
                ),
            )
            annexes.append(
                {
                    "file_name": evidence.get("file_name") or stored_name,
                    "description": evidence.get("description"),
                    "content": file_path.read_bytes(),
                }
            )
        except Exception:
            continue
    return annexes


def build_weg_wfr20997_cell_values(
    *,
    template_path: Path,
    plan: dict[str, Any],
    payload: dict[str, Any],
    team: list[dict[str, Any]],
    five_whys: dict[str, Any],
    actions: list[dict[str, Any]],
) -> dict[str, str]:
    anchors = load_merged_cell_anchors(template_path)
    values: dict[str, str] = {}

    def put(cell: str, value: object | None) -> None:
        put_cell_value(values, anchors, cell, value)

    def put_date(cell: str, value: str | date | datetime | None) -> None:
        put(cell, _format_date_value(value))

    nc = payload.get("nc_description") or {}
    classification = payload.get("classification") or {}
    effectiveness = payload.get("effectiveness") or {}
    preventive = payload.get("preventive") or {}
    containment_rows = payload.get("containment") or []
    documentation = payload.get("documentation_updates") or []

    branch = plan.get("branch_code") or "01"
    put("I4", plan.get("client_nc_registry"))
    put("E5", payload.get("supplier_name") or SUPPLIER_BY_BRANCH.get(branch, SUPPLIER_BY_BRANCH["01"]))
    put("E6", _material_label(plan))
    put("E7", payload.get("material_specification"))
    put("E8", payload.get("purchase_order"))
    put("E9", payload.get("invoice_number"))
    put_date("E10", payload.get("invoice_date"))
    put("E11", resolve_quantity_display(payload, "defective_quantity", "defective_quantity_unit"))
    put("E12", payload.get("return_invoice_number"))
    put("J5", resolve_delpi_primary_contact_name(plan))
    put("J6", resolve_delpi_contact_phone(plan))
    put("J7", payload.get("contact_fax"))
    put("J8", payload.get("client_batch"))
    put("J9", plan.get("batch_number"))
    put("J10", resolve_quantity_display(payload, "batch_quantity", "batch_quantity_unit"))
    put("J11", payload.get("disposition"))
    put("J12", resolve_quantity_display(payload, "rejected_quantity", "rejected_quantity_unit"))

    if classification.get("end_customer"):
        put("K3", plan.get("customer_name"))
    put_date("K1", payload.get("report_date") or plan.get("reported_at"))

    put("A15", nc.get("characteristic") or payload.get("nc_characteristic"))
    put("E15", nc.get("specified"))
    put("I15", nc.get("verified") or plan.get("reported_problem"))
    put("C18", nc.get("observations") or payload.get("observations"))

    put_date("D21", payload.get("return_by"))
    put("G21", resolve_customer_contact_name(plan))
    put("J21", resolve_customer_contact_email(plan))

    leader = next((member for member in team if member.get("is_leader")), team[0] if team else None)
    members = [member for member in team if not member.get("is_leader")]
    if leader:
        put("D23", leader.get("member_name"))
        put("H23", leader.get("department"))
    for index, member in enumerate(members[:4]):
        row = [25, 26, 27, 28][index]
        put(f"D{row}", member.get("member_name"))
        put(f"H{row}", member.get("department"))

    area_row_map = {
        "end_customer": 35,
        "client": 35,
        "client_plant": 37,
        "supplier": 39,
    }
    for item in containment_rows:
        area_key = (item.get("area") or "").lower()
        row = area_row_map.get(area_key)
        if not row:
            continue
        put(f"E{row}", item.get("quantity"))
        put(f"G{row}", item.get("action_plan"))
        put(f"J{row}", item.get("responsible"))
        put_date(f"M{row}", item.get("date"))

    occurrence_cols = ["E", "G", "I", "K", "M"]
    occurrence_whys = five_whys.get("occurrence_whys") or []
    for index, col in enumerate(occurrence_cols):
        raw = occurrence_whys[index] if index < len(occurrence_whys) else None
        put(f"{col}44", format_why_step_answer_cell(raw) if raw is not None else None)

    detection_cols = ["E", "G", "I", "K", "M"]
    detection_whys = five_whys.get("detection_whys") or []
    for index, col in enumerate(detection_cols):
        raw = detection_whys[index] if index < len(detection_whys) else None
        put(f"{col}49", format_why_step_answer_cell(raw) if raw is not None else None)

    corrective_actions = [
        action for action in actions if action.get("action_type") == "corrective" or action.get("cause_track")
    ]
    for index, action in enumerate(corrective_actions[:5]):
        row = [56, 57, 58, 59, 60][index]
        track = action.get("cause_track")
        if track == "occurrence":
            put(f"D{row}", "Ocorrência")
        elif track == "detection":
            put(f"D{row}", "Detecção")
        put(f"F{row}", action.get("description"))
        put(f"J{row}", _action_responsible_label(action))
        put_date(f"M{row}", action.get("due_date"))

    put("D63", effectiveness.get("resolved_how") or plan.get("effectiveness_notes"))
    put_date("D70", effectiveness.get("ok_material_date"))
    put("F70", effectiveness.get("new_parts_identification"))
    put("J70", effectiveness.get("verification_responsible"))
    put_date("L70", effectiveness.get("verification_date"))

    put("D72", preventive.get("how_avoid_future"))
    put("D77", preventive.get("other_processes_products"))
    put("D83", preventive.get("evaluation_responsible"))
    put_date("I83", preventive.get("evaluation_completion_date"))

    doc_rows = [85, 86, 87, 88]
    for index, doc in enumerate(documentation[:4]):
        row = doc_rows[index]
        put(f"F{row}", doc.get("document"))
        put(f"I{row}", doc.get("responsible"))
        put_date(f"K{row}", doc.get("date"))

    closure = payload.get("client_closure_note")
    if closure:
        put("D107", closure)
    return values


def _append_annex_images_to_workbook(
    workbook_bytes: bytes,
    image_annexes: list[dict[str, Any]],
) -> bytes:
    workbook = load_workbook(io.BytesIO(workbook_bytes))
    annex_ws = _resolve_annex_sheet(workbook)
    if annex_ws is None:
        return workbook_bytes

    _embed_annex_images(annex_ws, image_annexes)
    annex_buffer = io.BytesIO()
    workbook.save(annex_buffer)
    return overlay_zip_parts(
        base_bytes=workbook_bytes,
        overlay_bytes=annex_buffer.getvalue(),
        prefixes=(
            "xl/worksheets/sheet2.xml",
            "xl/worksheets/_rels/sheet2.xml.rels",
            "xl/drawings/drawing2",
        ),
    )


def build_rnc_8d_workbook(
    detail: dict[str, Any],
    *,
    image_annexes: list[dict[str, Any]] | None = None,
    template_key: str | None = None,
) -> bytes:
    plan = detail.get("plan") or {}
    resolved_key = resolve_export_template_key_for_plan(plan, requested_key=template_key)
    template_path = resolve_export_template_path(resolved_key)
    if not template_path.is_file():
        raise FileNotFoundError(f"Template 8D não encontrado: {template_path}")

    payload = plan.get("template_payload") or {}
    if not isinstance(payload, dict):
        payload = {}

    cell_values = build_weg_wfr20997_cell_values(
        template_path=template_path,
        plan=plan,
        payload=payload,
        team=detail.get("team_members") or [],
        five_whys=detail.get("five_whys") or {},
        actions=detail.get("actions") or [],
    )

    workbook_bytes = fill_template_workbook(template_path=template_path, cell_values=cell_values)
    if image_annexes:
        return _append_annex_images_to_workbook(workbook_bytes, image_annexes)
    return workbook_bytes
