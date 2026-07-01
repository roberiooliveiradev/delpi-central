from __future__ import annotations

import io
import re
import zipfile
from pathlib import Path

from openpyxl.utils import coordinate_to_tuple, get_column_letter

SHEET_DATA_PATTERN = re.compile(r"<sheetData\b[^>]*>.*?</sheetData>", re.DOTALL)
CELL_REF_PATTERN = re.compile(r'^<c r="(?P<ref>[A-Z]{1,3}[0-9]+)"')
CELL_OPEN_PATTERN = re.compile(r'^<c r="(?P<ref>[A-Z]{1,3}[0-9]+)"(?P<attrs>[^>/]*)(/?)>', re.DOTALL)
CELL_TYPE_PATTERN = re.compile(r'\bt="[^"]*"')
INVALID_XML_CHAR_PATTERN = re.compile(r"[\x00-\x08\x0B\x0C\x0E-\x1F\uFFFE\uFFFF]")
ROW_PATTERN_TEMPLATE = r'(<row\b[^>]*\br="{row}"[^>]*>)(.*?)(</row>)'

SHEET1_PATH = "xl/worksheets/sheet1.xml"
CALC_CHAIN_PATH = "xl/calcChain.xml"
CONTENT_TYPES_PATH = "[Content_Types].xml"
WORKBOOK_RELS_PATH = "xl/_rels/workbook.xml.rels"
CALC_CHAIN_CONTENT_TYPE = "/xl/calcChain.xml"
RELATIONSHIP_PATTERN = re.compile(
    r"<Relationship\b[^>]*Target=\"(?P<target>[^\"]+)\"[^>]*/>\s*",
    re.IGNORECASE,
)
CONTENT_TYPE_OVERRIDE_PATTERN = re.compile(
    r"<Override\b[^>]*PartName=\"(?P<part>[^\"]+)\"[^>]*/>\s*",
    re.IGNORECASE,
)
DRAWING_REFERENCE_PATTERN = re.compile(r"<drawing\s+r:id=\"[^\"]+\"\s*/>")
SHEET2_OVERLAY_PREFIXES = (
    "xl/worksheets/sheet2.xml",
    "xl/worksheets/_rels/sheet2.xml.rels",
    "xl/drawings/drawing2",
)


def _escape_xml_text(text: str) -> str:
    text = INVALID_XML_CHAR_PATTERN.sub("", text)
    return (
        text.replace("&", "&amp;")
        .replace("<", "&lt;")
        .replace(">", "&gt;")
        .replace("\r\n", "&#10;")
        .replace("\r", "&#10;")
        .replace("\n", "&#10;")
    )


def _parse_sheet_cells(sheet_data_xml: str) -> dict[str, str]:
    cells: dict[str, str] = {}
    cursor = 0
    while True:
        start = sheet_data_xml.find('<c r="', cursor)
        if start < 0:
            break
        self_closing = sheet_data_xml.find("/>", start)
        closing = sheet_data_xml.find("</c>", start)
        if self_closing >= 0 and (closing < 0 or self_closing < closing):
            end = self_closing + 2
        else:
            if closing < 0:
                break
            end = closing + 4
        chunk = sheet_data_xml[start:end]
        ref_match = CELL_REF_PATTERN.match(chunk)
        if ref_match is None:
            cursor = start + 4
            continue
        cells[ref_match.group("ref")] = chunk
        cursor = end
    return cells


def _inline_value_cell_xml(cell_xml: str, text: str) -> str:
    open_match = CELL_OPEN_PATTERN.match(cell_xml)
    if open_match is None:
        return cell_xml
    ref = open_match.group("ref")
    attrs = CELL_TYPE_PATTERN.sub("", open_match.group("attrs"))
    escaped = _escape_xml_text(text)
    return f'<c r="{ref}"{attrs} t="inlineStr"><is><t>{escaped}</t></is></c>'


def _row_number(cell_ref: str) -> int:
    return int(re.search(r"(\d+)", cell_ref).group(1))


def _insert_cell_in_sheet_data(sheet_data_xml: str, *, cell_ref: str, cell_xml: str) -> str:
    row_number = _row_number(cell_ref)
    row_pattern = re.compile(ROW_PATTERN_TEMPLATE.format(row=row_number), re.DOTALL)
    match = row_pattern.search(sheet_data_xml)
    if match is None:
        return sheet_data_xml + cell_xml
    updated_row = f"{match.group(1)}{match.group(2)}{cell_xml}{match.group(3)}"
    return sheet_data_xml.replace(match.group(0), updated_row, 1)


def patch_sheet_cell_values(sheet_xml: bytes, values: dict[str, str]) -> bytes:
    text = sheet_xml.decode("utf-8")
    sheet_data_match = SHEET_DATA_PATTERN.search(text)
    if sheet_data_match is None:
        return sheet_xml

    sheet_data = sheet_data_match.group(0)
    cells = _parse_sheet_cells(sheet_data)
    for cell_ref, value in values.items():
        if value is None or str(value).strip() == "":
            continue
        text_value = str(value).strip()
        existing = cells.get(cell_ref)
        if existing is not None:
            patched_cell = _inline_value_cell_xml(existing, text_value)
            sheet_data = sheet_data.replace(existing, patched_cell, 1)
            cells[cell_ref] = patched_cell
            continue
        new_cell = _inline_value_cell_xml(f'<c r="{cell_ref}"/>', text_value)
        sheet_data = _insert_cell_in_sheet_data(sheet_data, cell_ref=cell_ref, cell_xml=new_cell)

    patched = text.replace(sheet_data_match.group(0), sheet_data, 1)
    return patched.encode("utf-8")


def load_merged_cell_anchors(template_path: Path, *, sheet_name: str = "R8D") -> dict[str, str]:
    from openpyxl import load_workbook

    anchors: dict[str, str] = {}
    workbook = load_workbook(template_path, read_only=False, data_only=True)
    try:
        worksheet = workbook[sheet_name]
        for merged_range in worksheet.merged_cells.ranges:
            anchor = f"{get_column_letter(merged_range.min_col)}{merged_range.min_row}"
            for row in range(merged_range.min_row, merged_range.max_row + 1):
                for column in range(merged_range.min_col, merged_range.max_col + 1):
                    anchors[f"{get_column_letter(column)}{row}"] = anchor
    finally:
        workbook.close()
    return anchors


def resolve_template_cell(anchors: dict[str, str], cell: str) -> str:
    return anchors.get(cell, cell)


def _remove_relationships(rels_xml: bytes, *, targets: tuple[str, ...]) -> bytes:
    text = rels_xml.decode("utf-8")
    for target in targets:
        text = RELATIONSHIP_PATTERN.sub(
            lambda match: "" if match.group("target") == target else match.group(0),
            text,
        )
    return text.encode("utf-8")


def _remove_content_type_overrides(content_types: bytes, *, parts: tuple[str, ...]) -> bytes:
    text = content_types.decode("utf-8")
    for part in parts:
        text = CONTENT_TYPE_OVERRIDE_PATTERN.sub(
            lambda match: "" if match.group("part") == part else match.group(0),
            text,
        )
    return text.encode("utf-8")


def _sheet_has_drawing(sheet_xml: bytes) -> bool:
    return DRAWING_REFERENCE_PATTERN.search(sheet_xml.decode("utf-8")) is not None


def _strip_calc_chain_parts(patches: dict[str, bytes], skip_names: set[str]) -> None:
    skip_names.add(CALC_CHAIN_PATH)
    content_types = patches.get(CONTENT_TYPES_PATH)
    workbook_rels = patches.get(WORKBOOK_RELS_PATH)
    if content_types is None or workbook_rels is None:
        return
    patches[CONTENT_TYPES_PATH] = _remove_content_type_overrides(
        content_types,
        parts=(CALC_CHAIN_CONTENT_TYPE,),
    )
    patches[WORKBOOK_RELS_PATH] = _remove_relationships(
        workbook_rels,
        targets=(CALC_CHAIN_PATH,),
    )


def _write_zip_entries(
    *,
    source_zip: zipfile.ZipFile,
    output_zip: zipfile.ZipFile,
    patches: dict[str, bytes] | None = None,
    skip_names: set[str] | None = None,
) -> None:
    patches = patches or {}
    skip_names = skip_names or set()
    for item in source_zip.infolist():
        if item.filename in skip_names:
            continue
        data = patches.get(item.filename, source_zip.read(item.filename))
        output_zip.writestr(item, data)


def put_cell_value(
    values: dict[str, str],
    anchors: dict[str, str],
    cell: str,
    value: object | None,
) -> None:
    if value is None:
        return
    text = str(value).strip()
    if not text:
        return
    values[resolve_template_cell(anchors, cell)] = text


def fill_template_workbook(
    *,
    template_path: Path,
    cell_values: dict[str, str],
    sheet_path: str = SHEET1_PATH,
) -> bytes:
    buffer = io.BytesIO()
    with zipfile.ZipFile(template_path, "r") as template_zip:
        patches: dict[str, bytes] = {}
        skip_names: set[str] = set()
        if cell_values:
            patches[sheet_path] = patch_sheet_cell_values(
                template_zip.read(sheet_path),
                cell_values,
            )
            patches[CONTENT_TYPES_PATH] = template_zip.read(CONTENT_TYPES_PATH)
            patches[WORKBOOK_RELS_PATH] = template_zip.read(WORKBOOK_RELS_PATH)
            _strip_calc_chain_parts(patches, skip_names)
        with zipfile.ZipFile(buffer, "w") as output_zip:
            _write_zip_entries(
                source_zip=template_zip,
                output_zip=output_zip,
                patches=patches,
                skip_names=skip_names,
            )
    return buffer.getvalue()


def _extract_inline_string_cell_values(sheet_xml: bytes) -> dict[str, str]:
    sheet_data_match = SHEET_DATA_PATTERN.search(sheet_xml.decode("utf-8"))
    if sheet_data_match is None:
        return {}
    values: dict[str, str] = {}
    for ref, chunk in _parse_sheet_cells(sheet_data_match.group(0)).items():
        if 't="inlineStr"' not in chunk:
            continue
        text_match = re.search(r"<t[^>]*>(?P<text>.*?)</t>", chunk, re.DOTALL)
        if text_match is None:
            continue
        text = (
            text_match.group("text")
            .replace("&amp;", "&")
            .replace("&lt;", "<")
            .replace("&gt;", ">")
            .replace("&#10;", "\n")
            .strip()
        )
        if text:
            values[ref] = text
    return values


def overlay_zip_parts(*, base_bytes: bytes, overlay_bytes: bytes, prefixes: tuple[str, ...]) -> bytes:
    buffer = io.BytesIO()
    with zipfile.ZipFile(io.BytesIO(base_bytes), "r") as base_zip:
        overlay_names: set[str] = set()
        overlay_data: dict[str, bytes] = {}
        with zipfile.ZipFile(io.BytesIO(overlay_bytes), "r") as overlay_zip:
            for item in overlay_zip.infolist():
                if any(item.filename.startswith(prefix) for prefix in prefixes):
                    overlay_names.add(item.filename)
                    overlay_data[item.filename] = overlay_zip.read(item.filename)
                elif item.filename.startswith("xl/media/") and item.filename not in base_zip.namelist():
                    overlay_names.add(item.filename)
                    overlay_data[item.filename] = overlay_zip.read(item.filename)

        sheet2_path = "xl/worksheets/sheet2.xml"
        overlay_sheet2 = overlay_data.get(sheet2_path)
        base_sheet2 = base_zip.read(sheet2_path) if sheet2_path in base_zip.namelist() else None
        if (
            overlay_sheet2 is not None
            and base_sheet2 is not None
            and _sheet_has_drawing(base_sheet2)
            and not _sheet_has_drawing(overlay_sheet2)
        ):
            overlay_data[sheet2_path] = patch_sheet_cell_values(
                base_sheet2,
                _extract_inline_string_cell_values(overlay_sheet2),
            )

        with zipfile.ZipFile(buffer, "w") as output_zip:
            _write_zip_entries(
                source_zip=base_zip,
                output_zip=output_zip,
                patches={
                    name: data
                    for name, data in overlay_data.items()
                    if name in base_zip.namelist()
                },
            )
            with zipfile.ZipFile(io.BytesIO(overlay_bytes), "r") as overlay_zip:
                for item in overlay_zip.infolist():
                    if item.filename in overlay_names and item.filename not in base_zip.namelist():
                        output_zip.writestr(item, overlay_data[item.filename])
    return buffer.getvalue()


def patch_workbook_sheet(
    workbook_bytes: bytes,
    *,
    sheet_path: str,
    cell_values: dict[str, str],
) -> bytes:
    buffer = io.BytesIO()
    with zipfile.ZipFile(io.BytesIO(workbook_bytes), "r") as source_zip:
        patches: dict[str, bytes] = {}
        skip_names: set[str] = set()
        if cell_values:
            patches[sheet_path] = patch_sheet_cell_values(
                source_zip.read(sheet_path),
                cell_values,
            )
            patches[CONTENT_TYPES_PATH] = source_zip.read(CONTENT_TYPES_PATH)
            patches[WORKBOOK_RELS_PATH] = source_zip.read(WORKBOOK_RELS_PATH)
            _strip_calc_chain_parts(patches, skip_names)
        with zipfile.ZipFile(buffer, "w") as output_zip:
            _write_zip_entries(
                source_zip=source_zip,
                output_zip=output_zip,
                patches=patches,
                skip_names=skip_names,
            )
    return buffer.getvalue()
