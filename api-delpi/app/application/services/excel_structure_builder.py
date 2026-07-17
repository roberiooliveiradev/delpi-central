# app/application/services/excel_structure_builder.py

import io
from string import ascii_uppercase

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from app.domain.entities.product.bom_node import BomNode


class ExcelStructureBuilder:

    @staticmethod
    def build(root: BomNode) -> io.BytesIO:
        """
        Gera planilha Excel no padrão DELPI a partir da árvore BOM.
        """

        rows: list[list] = []
        meta_map: dict[str, dict] = {}

        # -------------------------------------------------
        # Cabeçalho do produto pai (se não houver MP direta)
        # -------------------------------------------------

        has_direct_mp = any(c.type == "MP" for c in root.components)

        if not has_direct_mp:
            rows.append([
                root.code,
                root.description,
                "",
                "",
                "",
                "",
                root.type,
                root.unit,
                1.0
            ])

        # -------------------------------------------------
        # Recursão com parent_factor acumulativo
        # -------------------------------------------------

        def process_components(
            parent_code: str,
            parent_desc: str,
            components: list[BomNode],
            parent_factor: float = 1.0
        ):

            for comp in components:

                code_comp = comp.code
                desc_comp = comp.description
                comp_type = comp.type
                comp_unit = comp.unit
                comp_qtd = float(comp.quantity or 1)

                if comp_type == "MP":

                    rows.append([
                        parent_code,
                        parent_desc,
                        "",
                        comp_qtd,
                        code_comp,
                        desc_comp,
                        comp_type,
                        comp_unit,
                        parent_factor,
                    ])

                    meta_map[code_comp] = {
                        "type": comp_type,
                        "unit": comp_unit
                    }

                else:

                    new_parent_factor = parent_factor * comp_qtd

                    if comp.components:
                        process_components(
                            code_comp,
                            desc_comp,
                            comp.components,
                            new_parent_factor
                        )
                    else:
                        # Intermediário sem MP amarrada: listar como componente
                        # do pai — senão o item some da planilha.
                        rows.append([
                            parent_code,
                            parent_desc,
                            "",
                            comp_qtd,
                            code_comp,
                            desc_comp,
                            comp_type,
                            comp_unit,
                            parent_factor,
                        ])

                        meta_map[code_comp] = {
                            "type": comp_type,
                            "unit": comp_unit
                        }

        if root.components:
            process_components(
                root.code,
                root.description,
                root.components,
                1.0
            )

        # -------------------------------------------------
        # Gerar coluna Item (A, B, C...)
        # -------------------------------------------------

        def generate_item_label(index: int) -> str:

            letters = ""

            while True:
                index, remainder = divmod(index, 26)
                letters = ascii_uppercase[remainder] + letters

                if index == 0:
                    break

                index -= 1

            return letters

        item_map: dict[str, str] = {}
        item_counter = 0

        for r in rows:

            comp_code = r[4]

            if comp_code:

                if comp_code not in item_map:
                    item_map[comp_code] = generate_item_label(item_counter)
                    item_counter += 1

                r[2] = item_map[comp_code]

        # -------------------------------------------------
        # Criar planilha
        # -------------------------------------------------

        wb = Workbook()
        ws = wb.active
        ws.title = "Estrutura DELPI"

        headers = ["Código", "Descrição", "Item", "QTD", "Componente", "Descrição"]
        ws.append(headers)

        font_name = "Arial Narrow"
        font_size = 10

        thin = Side(border_style="thin", color="000000")

        border = Border(
            left=thin,
            right=thin,
            top=thin,
            bottom=thin
        )

        align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
        align_left = Alignment(horizontal="left", vertical="center", wrap_text=True)

        header_font = Font(bold=True, color="0000FF", name=font_name, size=font_size)
        normal_font = Font(color="000000", name=font_name, size=font_size)
        red_font = Font(color="FF0000", name=font_name, size=font_size)
        blue_font = Font(color="0000FF", name=font_name, size=font_size)

        for col in range(1, len(headers) + 1):

            c = ws.cell(row=1, column=col)

            c.font = header_font
            c.alignment = align_center
            c.border = border

            ws.column_dimensions[get_column_letter(col)].width = 50 if col in [2, 6] else 14

        for r in rows:
            ws.append(r[:6])

        # -------------------------------------------------
        # Aplicar formatação base
        # -------------------------------------------------

        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=len(headers)):

            for c in row:

                c.border = border

                if c.column == 6:
                    c.alignment = align_left
                else:
                    c.alignment = align_center

                c.font = normal_font

        # -------------------------------------------------
        # Mesclar Código / Descrição
        # -------------------------------------------------

        current_parent = None
        start_row = 2

        for i in range(2, ws.max_row + 1):

            parent = ws.cell(row=i, column=1).value

            if parent != current_parent:

                if current_parent is not None and start_row < i - 1:

                    ws.merge_cells(start_row=start_row, start_column=1, end_row=i - 1, end_column=1)
                    ws.merge_cells(start_row=start_row, start_column=2, end_row=i - 1, end_column=2)

                start_row = i
                current_parent = parent

        if current_parent and start_row < ws.max_row:

            ws.merge_cells(start_row=start_row, start_column=1, end_row=ws.max_row, end_column=1)
            ws.merge_cells(start_row=start_row, start_column=2, end_row=ws.max_row, end_column=2)

        # -------------------------------------------------
        # Ajuste final de quantidade
        # -------------------------------------------------

        for idx, row in enumerate(
            ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=len(headers)),
            start=0
        ):

            qtd_cell = row[3].value
            comp_code = str(row[4].value or "").strip()

            comp_meta = meta_map.get(comp_code)

            if not comp_meta:

                for c in row:
                    c.font = blue_font

                continue

            comp_unit = comp_meta["unit"]

            try:
                qtd = float(str(qtd_cell).replace(",", "."))
            except Exception:
                qtd = 0

            parent_factor = rows[idx][8] if len(rows[idx]) > 8 else 1

            if comp_unit == "PC":

                if qtd % 1000 == 0:
                    qtd = qtd / 1000
                else:
                    qtd = 1

            else:
                qtd = 1

            qtd_final = qtd * parent_factor

            row[3].value = qtd_final

            for c in row:
                c.font = blue_font

            if qtd_final >= 2:

                for c in row[2:6]:
                    c.font = red_font

        # -------------------------------------------------
        # Retorno
        # -------------------------------------------------

        stream = io.BytesIO()

        wb.save(stream)

        stream.seek(0)

        return stream