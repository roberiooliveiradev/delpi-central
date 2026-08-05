"""Geração Excel da consolidação gerencial CAPEX (Fase 2D.1)."""

from __future__ import annotations

import io
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from app.domain.services.planejamento_orcamentario.capex_consolidation_constants import (
    PLAN_STATUS_LABELS,
)


def _header_style(ws: Worksheet, row: int = 1) -> None:
    font = Font(bold=True)
    for cell in ws[row]:
        cell.font = font
        cell.alignment = Alignment(wrap_text=True, vertical="center")
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions


def _autosize(ws: Worksheet, min_width: int = 12, max_width: int = 42) -> None:
    for idx, column_cells in enumerate(ws.columns, start=1):
        length = 0
        for cell in column_cells:
            value = "" if cell.value is None else str(cell.value)
            length = max(length, len(value))
        ws.column_dimensions[get_column_letter(idx)].width = min(
            max(length + 2, min_width), max_width
        )


def _as_number(value: Any) -> float | None:
    if value is None or value == "":
        return None
    try:
        return float(Decimal(str(value)))
    except (InvalidOperation, ValueError, TypeError):
        return None


def _as_date(value: Any) -> date | None:
    if value is None or value == "":
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    text = str(value)[:10]
    try:
        return date.fromisoformat(text)
    except ValueError:
        return None


def _write_kv(ws: Worksheet, rows: list[tuple[str, Any]], start_row: int = 1) -> int:
    r = start_row
    for label, value in rows:
        ws.cell(row=r, column=1, value=label)
        ws.cell(row=r, column=2, value=value)
        r += 1
    return r


def build_capex_consolidation_workbook(
    *,
    exercise: dict[str, Any],
    filters: dict[str, Any],
    summary: dict[str, Any],
    details: list[dict[str, Any]],
    by_cost_center: list[dict[str, Any]],
    by_category: list[dict[str, Any]],
    by_month: list[dict[str, Any]],
    generated_at: datetime,
    generated_by: str,
) -> io.BytesIO:
    wb = Workbook()

    # --- Resumo ---
    ws_summary = wb.active
    ws_summary.title = "Resumo"
    thin = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    year = exercise.get("year")
    filter_lines = []
    for key, value in filters.items():
        if value is None or value == "" or value == []:
            continue
        filter_lines.append(f"{key}={value}")
    next_row = _write_kv(
        ws_summary,
        [
            ("Exercício", f"{year} — {exercise.get('name')}" if year else str(exercise.get("id"))),
            ("Exercício ID", str(exercise.get("id") or "")),
            ("Gerado em", generated_at.isoformat(timespec="seconds")),
            ("Gerado por", generated_by),
            ("Filtros aplicados", "; ".join(filter_lines) if filter_lines else "(nenhum além do exercício)"),
            ("Moeda", summary.get("currency")),
            ("Valor total previsto", _as_number(summary.get("total_estimated_amount"))),
            ("Quantidade de investimentos", summary.get("investment_count")),
            ("Quantidade de centros de custo", summary.get("cost_center_count")),
            ("Planos em rascunho", summary.get("plans_draft_count")),
            ("Planos enviados", summary.get("plans_submitted_count")),
            ("Planos com ajustes solicitados", summary.get("plans_changes_requested_count")),
            ("Planos reprovados", summary.get("plans_rejected_count")),
            ("Planos aprovados", summary.get("plans_approved_count")),
            ("Valor aprovado", _as_number(summary.get("approved_amount"))),
            ("Valor em análise", _as_number(summary.get("in_review_amount"))),
            ("Investimentos incompletos", summary.get("incomplete_investment_count")),
        ],
    )
    for row in ws_summary.iter_rows(min_row=1, max_row=next_row - 1, max_col=2):
        for cell in row:
            cell.border = thin
        if isinstance(row[1].value, (int, float)) and row[0].value and "Valor" in str(row[0].value):
            row[1].number_format = '#,##0.00'
    ws_summary.column_dimensions["A"].width = 36
    ws_summary.column_dimensions["B"].width = 48

    # --- Investimentos ---
    ws_inv = wb.create_sheet("Investimentos")
    inv_headers = [
        "Exercício",
        "Unidade",
        "Área",
        "Centro de custo",
        "Responsável",
        "Descrição",
        "Categoria",
        "Prioridade",
        "Origem",
        "Fornecedor provável",
        "Valor previsto",
        "Moeda",
        "Data Rcbto",
        "Completo",
        "Status do planejamento",
    ]
    ws_inv.append(inv_headers)
    for item in details:
        ws_inv.append(
            [
                str(item.get("exercise_id") or ""),
                item.get("unit_id"),
                item.get("area_id"),
                item.get("cost_center_id"),
                item.get("responsible"),
                item.get("description"),
                item.get("category_name") or item.get("category_code") or item.get("category_id"),
                item.get("priority_label") or item.get("priority"),
                item.get("origin_label") or item.get("origin"),
                item.get("probable_supplier_name"),
                _as_number(item.get("estimated_amount")),
                item.get("currency"),
                _as_date(item.get("required_date")),
                "Sim" if item.get("is_complete") else "Não",
                item.get("plan_status_label")
                or PLAN_STATUS_LABELS.get(str(item.get("plan_status") or ""), item.get("plan_status")),
            ]
        )
    _header_style(ws_inv)
    for row in ws_inv.iter_rows(min_row=2, min_col=11, max_col=11):
        for cell in row:
            if isinstance(cell.value, (int, float)):
                cell.number_format = '#,##0.00'
    for row in ws_inv.iter_rows(min_row=2, min_col=13, max_col=13):
        for cell in row:
            if isinstance(cell.value, date):
                cell.number_format = "YYYY-MM-DD"
    _autosize(ws_inv)

    # --- Por Centro de Custo ---
    ws_cc = wb.create_sheet("Por Centro de Custo")
    ws_cc.append(
        ["Unidade", "Área", "Centro de custo", "Quantidade", "Valor total", "Status do plano"]
    )
    for item in by_cost_center:
        ws_cc.append(
            [
                item.get("unit_id"),
                item.get("area_id"),
                item.get("cost_center_id") or item.get("code"),
                item.get("investment_count"),
                _as_number(item.get("total_amount")),
                item.get("plan_status_label") or item.get("plan_status"),
            ]
        )
    _header_style(ws_cc)
    for row in ws_cc.iter_rows(min_row=2, min_col=5, max_col=5):
        for cell in row:
            if isinstance(cell.value, (int, float)):
                cell.number_format = '#,##0.00'
    _autosize(ws_cc)

    # --- Por Categoria ---
    ws_cat = wb.create_sheet("Por Categoria")
    ws_cat.append(["Categoria", "Quantidade", "Valor total", "Percentual"])
    for item in by_category:
        pct = item.get("percent_of_total")
        ws_cat.append(
            [
                item.get("description") or item.get("code"),
                item.get("investment_count"),
                _as_number(item.get("total_amount")),
                None if pct is None else float(Decimal(str(pct))) / 100.0,
            ]
        )
    _header_style(ws_cat)
    for row in ws_cat.iter_rows(min_row=2, min_col=3, max_col=3):
        for cell in row:
            if isinstance(cell.value, (int, float)):
                cell.number_format = '#,##0.00'
    for row in ws_cat.iter_rows(min_row=2, min_col=4, max_col=4):
        for cell in row:
            if isinstance(cell.value, (int, float)):
                cell.number_format = "0.00%"
    _autosize(ws_cat)

    # --- Por Mês ---
    ws_month = wb.create_sheet("Por Mês")
    ws_month.append(["Mês da Data Rcbto", "Quantidade", "Valor total"])
    for item in by_month:
        ws_month.append(
            [
                item.get("description") or item.get("code"),
                item.get("investment_count"),
                _as_number(item.get("total_amount")),
            ]
        )
    _header_style(ws_month)
    for row in ws_month.iter_rows(min_row=2, min_col=3, max_col=3):
        for cell in row:
            if isinstance(cell.value, (int, float)):
                cell.number_format = '#,##0.00'
    _autosize(ws_month)

    stream = io.BytesIO()
    wb.save(stream)
    stream.seek(0)
    return stream
