#!/usr/bin/env python3
"""Gera fixtures mínimas e garante templates de export 8D no build."""

from __future__ import annotations

import shutil
from pathlib import Path

from openpyxl import Workbook

ROOT = Path(__file__).resolve().parents[1]
TEMPLATES_DIR = ROOT / "app" / "content" / "templates" / "quality"
FIXTURE = ROOT / "tests" / "fixtures" / "quality" / "rnc_8d_template_minimal.xlsx"
WEG = TEMPLATES_DIR / "weg_wfr20997_template.xlsx"
DELPI = TEMPLATES_DIR / "delpi_8d_template.xlsx"


def _write_minimal(path: Path) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = "R8D"
    ws["I4"] = ""
    annex = wb.create_sheet("Anexos(Evidencias)")
    annex["A1"] = "Evidências"
    wb.save(path)


def _ensure_delpi_branding(path: Path) -> None:
    if not path.is_file():
        return
    from openpyxl import load_workbook

    wb = load_workbook(path)
    ws = wb["R8D"]
    ws["C1"] = 1
    ws["D1"] = "RELATÓRIO 8D — NÃO CONFORMIDADE (DELPI)"
    ws["A110"] = "DELPI Componentes — PAC Qualidade"
    wb.save(path)


def main() -> int:
    TEMPLATES_DIR.mkdir(parents=True, exist_ok=True)
    _write_minimal(FIXTURE)

    if not WEG.is_file():
        _write_minimal(WEG)
        print(f"[OK] Template WEG mínimo gerado: {WEG}")
    else:
        print(f"[OK] Template WEG existente: {WEG}")

    if not DELPI.is_file():
        shutil.copy2(WEG, DELPI)
        _ensure_delpi_branding(DELPI)
        print(f"[OK] Template DELPI gerado a partir do WEG: {DELPI}")
    else:
        _ensure_delpi_branding(DELPI)
        print(f"[OK] Template DELPI existente: {DELPI}")

    print(f"[OK] Fixture mínima: {FIXTURE}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
