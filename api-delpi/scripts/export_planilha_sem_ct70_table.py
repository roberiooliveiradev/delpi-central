"""Gera tabela: produtos na planilha sem CT-70 × apontamentos Totvs (outros CTs)."""
from __future__ import annotations

import csv
from collections import defaultdict
from dataclasses import dataclass
from datetime import date, datetime, timedelta
from pathlib import Path

import openpyxl

from app.infrastructure.persistence.totvs.ppm_repositories.ppm_query_repository import (
    PpmQueryRepository,
)

SHEET_PATH = (
    Path(__file__).resolve().parents[1]
    / "planilhas"
    / "RQ 005 Inspeção da Qualidade Rev08 Matriz - 2026.xlsm"
)
OUT_DIR = Path(__file__).resolve().parents[1] / "output"
OUT_MD = OUT_DIR / "tabela_planilha_sem_ct70_apontamentos_totvs.md"
OUT_CSV = OUT_DIR / "tabela_planilha_sem_ct70_apontamentos_totvs.csv"

MONTHS = [
    ("Jan", "2026-01-01", "2026-01-31"),
    ("Fev", "2026-02-01", "2026-02-28"),
    ("Mar", "2026-03-01", "2026-03-31"),
    ("Abr", "2026-04-01", "2026-04-30"),
    ("Mai", "2026-05-01", "2026-05-31"),
]

LOT_COL_START = 8
OP_ROW = 2
PRODUCT_ROW = 3
QTY_ROW = 4
DATE_ROW = 8

SQL_CT70_BY_PRODUCT = """
WITH ct_inspecao_final AS (
    SELECT HB.HB_FILIAL, HB.HB_COD AS ct_inspecao
    FROM SHB010 HB
    WHERE HB.D_E_L_E_T_ = ' '
      AND UPPER(HB.HB_NOME) LIKE '%INSPE%FINAL%'
      AND HB.HB_FILIAL = '01'
)
SELECT
    LTRIM(RTRIM(SH6.H6_PRODUTO)) AS product_code,
    SUM(CAST(SH6.H6_QTDPROD AS FLOAT)) * 1000 AS total_un
FROM SH6010 SH6
INNER JOIN SH1010 SH1
    ON SH1.H1_FILIAL = SH6.H6_FILIAL
   AND SH1.H1_CODIGO = SH6.H6_RECURSO
   AND SH1.D_E_L_E_T_ = ' '
INNER JOIN ct_inspecao_final CIF
    ON CIF.HB_FILIAL = SH6.H6_FILIAL
   AND CIF.ct_inspecao = SH1.H1_CTRAB
WHERE SH6.D_E_L_E_T_ = ' '
  AND SH6.H6_FILIAL = '01'
  AND SH6.H6_TIPO = 'P'
  AND SH6.H6_DTAPONT >= '20260101'
  AND SH6.H6_DTAPONT < '20260601'
GROUP BY SH6.H6_PRODUTO
"""

SQL_APONT_BY_OPS = """
SELECT
    LTRIM(RTRIM(SH6.H6_OP)) AS op,
    LTRIM(RTRIM(SH6.H6_PRODUTO)) AS product_code,
    LTRIM(RTRIM(SH1.H1_CTRAB)) AS ct,
    LTRIM(RTRIM(HB.HB_NOME)) AS ct_nome,
    SH6.H6_DTAPONT AS dt_apont,
    CAST(SH6.H6_QTDPROD AS FLOAT) * 1000 AS qtd_un
FROM SH6010 SH6
LEFT JOIN SH1010 SH1
    ON SH1.H1_FILIAL = SH6.H6_FILIAL
   AND SH1.H1_CODIGO = SH6.H6_RECURSO
   AND SH1.D_E_L_E_T_ = ' '
LEFT JOIN SHB010 HB
    ON HB.HB_FILIAL = SH6.H6_FILIAL
   AND HB.HB_COD = SH1.H1_CTRAB
   AND HB.D_E_L_E_T_ = ' '
WHERE SH6.D_E_L_E_T_ = ' '
  AND SH6.H6_FILIAL = '01'
  AND SH6.H6_TIPO = 'P'
  AND SH6.H6_OP IN ({placeholders})
ORDER BY SH6.H6_OP, SH6.H6_DTAPONT, SH1.H1_CTRAB
"""


@dataclass
class PlanilhaLot:
    month: str
    col: str
    op: str
    product_code: str
    qty: float
    plan_date: str


@dataclass
class TableRow:
    month: str
    col: str
    product_code: str
    op: str
    qty_planilha: float
    date_planilha: str
    ct_totvs: str
    ct_nome: str
    date_totvs: str
    qty_totvs: float
    apontado_ct70: str


def _format_plan_date(raw) -> str:
    if raw is None:
        return ""
    if isinstance(raw, datetime):
        return raw.strftime("%Y-%m-%d")
    if isinstance(raw, date):
        return raw.isoformat()
    text = str(raw).strip()
    if len(text) == 8 and text.isdigit():
        return f"{text[0:4]}-{text[4:6]}-{text[6:8]}"
    return text


def _format_totvs_date(raw) -> str:
    text = str(raw or "").strip()
    if len(text) == 8 and text.isdigit():
        return f"{text[0:4]}-{text[4:6]}-{text[6:8]}"
    return text


def _normalize_product(raw) -> str | None:
    if raw is None:
        return None
    code = str(raw).strip().replace(".0", "")
    if not code or code.lower().startswith("código"):
        return None
    return code


def _normalize_qty(raw) -> float:
    if raw is None:
        return 0.0
    try:
        return float(raw)
    except (TypeError, ValueError):
        return 0.0


def parse_planilha_lots(path: Path) -> list[PlanilhaLot]:
    lots: list[PlanilhaLot] = []
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True, keep_links=False)
    month_names = [name for name, _, _ in MONTHS if name in wb.sheetnames]

    for sheet_name in month_names:
        ws = wb[sheet_name]
        op_row: tuple = ()
        product_row: tuple = ()
        qty_row: tuple = ()
        date_row: tuple = ()

        for row_idx, row in enumerate(
            ws.iter_rows(min_row=1, max_row=DATE_ROW, values_only=True),
            start=1,
        ):
            if row_idx == OP_ROW:
                op_row = row
            elif row_idx == PRODUCT_ROW:
                product_row = row
            elif row_idx == QTY_ROW:
                qty_row = row
            elif row_idx == DATE_ROW:
                date_row = row

        max_len = max(len(op_row), len(product_row), len(qty_row), len(date_row))
        for col_idx in range(LOT_COL_START - 1, max_len):
            product = _normalize_product(
                product_row[col_idx] if col_idx < len(product_row) else None
            )
            qty = _normalize_qty(qty_row[col_idx] if col_idx < len(qty_row) else None)
            if not product or qty <= 0:
                continue

            op_raw = op_row[col_idx] if col_idx < len(op_row) else None
            op = str(op_raw).strip().replace(".0", "") if op_raw else ""
            plan_date = _format_plan_date(
                date_row[col_idx] if col_idx < len(date_row) else None
            )
            col_letter = openpyxl.utils.get_column_letter(col_idx + 1)
            lots.append(
                PlanilhaLot(
                    month=sheet_name,
                    col=col_letter,
                    op=op,
                    product_code=product,
                    qty=qty,
                    plan_date=plan_date,
                )
            )

    wb.close()
    return lots


def fetch_zero_ct70_products() -> set[str]:
    with PpmQueryRepository() as repo:
        rows = repo.execute_query(SQL_CT70_BY_PRODUCT, ()) or []
    return {
        str(row.get("product_code") or "").strip()
        for row in rows
        if float(row.get("total_un") or 0) <= 0
    }


def fetch_ct70_totals() -> dict[str, float]:
    with PpmQueryRepository() as repo:
        rows = repo.execute_query(SQL_CT70_BY_PRODUCT, ()) or []
    totals: dict[str, float] = defaultdict(float)
    for row in rows:
        code = str(row.get("product_code") or "").strip()
        totals[code] = float(row.get("total_un") or 0)
    return dict(totals)


def fetch_apontamentos_by_ops(ops: list[str]) -> dict[str, list[dict]]:
    if not ops:
        return {}

    chunk_size = 200
    by_op: dict[str, list[dict]] = defaultdict(list)

    with PpmQueryRepository() as repo:
        for i in range(0, len(ops), chunk_size):
            chunk = ops[i : i + chunk_size]
            placeholders = ", ".join("?" for _ in chunk)
            sql = SQL_APONT_BY_OPS.format(placeholders=placeholders)
            rows = repo.execute_query(sql, tuple(chunk)) or []
            for row in rows:
                op = str(row.get("op") or "").strip()
                by_op[op].append(row)

    return dict(by_op)


def build_table_rows(
    lots: list[PlanilhaLot],
    zero_ct70: set[str],
) -> list[TableRow]:
    target_lots = [lot for lot in lots if lot.product_code in zero_ct70]
    ops = sorted({lot.op for lot in target_lots if lot.op})
    apont_by_op = fetch_apontamentos_by_ops(ops)

    ct70_names = {"CT-70", "CT-70 "}

    rows: list[TableRow] = []
    for lot in target_lots:
        aponts = apont_by_op.get(lot.op, [])
        non_ct70 = [
            a
            for a in aponts
            if str(a.get("ct") or "").strip() not in ct70_names
            and "INSPE" not in str(a.get("ct_nome") or "").upper()
        ]

        if not aponts:
            rows.append(
                TableRow(
                    month=lot.month,
                    col=lot.col,
                    product_code=lot.product_code,
                    op=lot.op,
                    qty_planilha=lot.qty,
                    date_planilha=lot.plan_date,
                    ct_totvs="—",
                    ct_nome="Sem apontamento no Totvs",
                    date_totvs="—",
                    qty_totvs=0.0,
                    apontado_ct70="Não",
                )
            )
            continue

        if not non_ct70:
            # OP existe, mas só CT inspeção (improvável para zero_ct70 product)
            for ap in aponts:
                rows.append(
                    TableRow(
                        month=lot.month,
                        col=lot.col,
                        product_code=lot.product_code,
                        op=lot.op,
                        qty_planilha=lot.qty,
                        date_planilha=lot.plan_date,
                        ct_totvs=str(ap.get("ct") or "—"),
                        ct_nome=str(ap.get("ct_nome") or ""),
                        date_totvs=_format_totvs_date(ap.get("dt_apont")),
                        qty_totvs=float(ap.get("qtd_un") or 0),
                        apontado_ct70="Não",
                    )
                )
            continue

        matched = [
            ap
            for ap in non_ct70
            if str(ap.get("product_code") or "").strip() == lot.product_code
        ] or non_ct70

        for ap in matched:
            rows.append(
                TableRow(
                    month=lot.month,
                    col=lot.col,
                    product_code=lot.product_code,
                    op=lot.op,
                    qty_planilha=lot.qty,
                    date_planilha=lot.plan_date,
                    ct_totvs=str(ap.get("ct") or "—"),
                    ct_nome=str(ap.get("ct_nome") or ""),
                    date_totvs=_format_totvs_date(ap.get("dt_apont")),
                    qty_totvs=float(ap.get("qtd_un") or 0),
                    apontado_ct70="Não",
                )
            )

    return rows


def write_csv(rows: list[TableRow]) -> None:
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    with OUT_CSV.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.writer(handle)
        writer.writerow(
            [
                "mes",
                "coluna",
                "codigo_produto",
                "op",
                "qtd_planilha",
                "data_planilha",
                "ct_totvs",
                "nome_ct",
                "data_apont_totvs",
                "qtd_apont_totvs",
                "apontado_ct70",
            ]
        )
        for row in rows:
            writer.writerow(
                [
                    row.month,
                    row.col,
                    row.product_code,
                    row.op,
                    f"{row.qty_planilha:.0f}",
                    row.date_planilha,
                    row.ct_totvs,
                    row.ct_nome,
                    row.date_totvs,
                    f"{row.qty_totvs:.0f}",
                    row.apontado_ct70,
                ]
            )


def write_markdown(rows: list[TableRow], zero_products: list[str]) -> None:
    lot_count = len({(r.month, r.col, r.op, r.product_code) for r in rows})
    apont_count = len(rows)

    lines = [
        "# Planilha RQ 005 — produtos sem apontamento no CT-70",
        "",
        "**Filial:** 01 (Matriz)  ",
        "**Período:** jan–mai/2026  ",
        "**Gerado em:** "
        + datetime.now().strftime("%d/%m/%Y"),
        "",
        "Produtos com **zero** apontamento no CT de inspeção final no período, "
        "cruzados com apontamentos reais no Protheus (outros CTs) e data da planilha (linha 8).",
        "",
        f"**{len(zero_products)} códigos** · **{lot_count} lotes na planilha** · "
        f"**{apont_count} linhas** (uma por apontamento Totvs vinculado à OP do lote).",
        "",
        "CSV completo: [`tabela_planilha_sem_ct70_apontamentos_totvs.csv`](./tabela_planilha_sem_ct70_apontamentos_totvs.csv)",
        "",
        "## Códigos sem CT-70",
        "",
        "| Código | Observação |",
        "|--------|------------|",
    ]

    notes = {
        "50232465": "PI — apontado em montagem/corte, não passa pelo CT-70",
        "50233615": "PI — idem",
        "50233616": "PI — idem",
        "550232465": "Typo provável de 50232465 (Jan!LX) — sem cadastro/apontamento",
        "9023826": "Código inexistente no SB1 (Mai — provável 90263826)",
        "90151384": "Código inexistente no SB1",
        "90252552": "Código inexistente no SB1",
        "990263749": "Código inexistente no SB1",
    }
    for code in zero_products:
        lines.append(f"| {code} | {notes.get(code, 'Sem apontamento CT-70')} |")

    lines.extend(
        [
            "",
            "## Tabela detalhada (amostra — primeiros 80 apontamentos)",
            "",
            "| Mês | Col | Produto | OP | Qtd plan. | Data plan. | CT Totvs | Data Totvs | Qtd Totvs |",
            "|-----|-----|---------|-----|----------:|------------|----------|------------|----------:|",
        ]
    )

    for row in rows[:80]:
        ct_label = row.ct_totvs
        if row.ct_nome and row.ct_nome != "Sem apontamento no Totvs":
            ct_label = f"{row.ct_totvs} ({row.ct_nome[:28]})"
        lines.append(
            f"| {row.month} | {row.col} | {row.product_code} | {row.op} | "
            f"{row.qty_planilha:,.0f} | {row.date_planilha} | {ct_label} | "
            f"{row.date_totvs} | {row.qty_totvs:,.0f} |"
        )

    if len(rows) > 80:
        lines.append("")
        lines.append(
            f"*… mais {len(rows) - 80} linhas no CSV "
            f"([`tabela_planilha_sem_ct70_apontamentos_totvs.csv`](./tabela_planilha_sem_ct70_apontamentos_totvs.csv)).*"
        )

    # Resumo por produto + CT
    lines.extend(["", "## Resumo por produto e CT Totvs", ""])
    agg: dict[tuple[str, str], float] = defaultdict(float)
    for row in rows:
        if row.ct_totvs == "—":
            continue
        agg[(row.product_code, row.ct_totvs)] += row.qty_totvs

    lines.append("| Produto | CT Totvs | Qtd apontada (un.) |")
    lines.append("|---------|----------|-------------------:|")
    for (product, ct), qty in sorted(agg.items(), key=lambda x: (-x[1], x[0][0])):
        lines.append(f"| {product} | {ct} | {qty:,.0f} |")

    OUT_MD.write_text("\n".join(lines) + "\n", encoding="utf-8")


def main() -> None:
    if not SHEET_PATH.exists():
        raise SystemExit(f"Planilha não encontrada: {SHEET_PATH}")

    lots = parse_planilha_lots(SHEET_PATH)
    ct70_totals = fetch_ct70_totals()
    all_products = {lot.product_code for lot in lots}
    zero_ct70 = {code for code in all_products if ct70_totals.get(code, 0) <= 0}
    zero_products = sorted(
        zero_ct70,
        key=lambda c: -sum(lot.qty for lot in lots if lot.product_code == c),
    )

    rows = build_table_rows(lots, zero_ct70)
    write_csv(rows)
    write_markdown(rows, zero_products)

    print(f"Produtos sem CT-70: {len(zero_products)}")
    print(f"Lotes planilha: {len({(r.month, r.col, r.op, r.product_code) for r in rows})}")
    print(f"Linhas tabela: {len(rows)}")
    print(f"Markdown: {OUT_MD}")
    print(f"CSV: {OUT_CSV}")


if __name__ == "__main__":
    main()
