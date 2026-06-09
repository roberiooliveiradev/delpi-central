"""Cruza produtos da planilha RQ 005 (linha 4) com apontamentos no CT-70 (filial 01)."""
from __future__ import annotations

from collections import defaultdict
from dataclasses import dataclass, field
from datetime import datetime, timedelta
from pathlib import Path

import openpyxl

from app.infrastructure.persistence.totvs.ppm_repositories.ppm_query_repository import (
    PpmQueryRepository,
)

SHEET_PATH = (
    Path(__file__).resolve().parents[1]
    / "planilhas"
    / "RQ 005 Inspeção da Qualidade Rev08 Matriz - 2026.xlsm"
)

MONTHS = [
    ("Jan", "2026-01-01", "2026-01-31"),
    ("Fev", "2026-02-01", "2026-02-28"),
    ("Mar", "2026-03-01", "2026-03-31"),
    ("Abr", "2026-04-01", "2026-04-30"),
    ("Mai", "2026-05-01", "2026-05-31"),
]

# Colunas H em diante = lotes (col 8+)
LOT_COL_START = 8
PRODUCT_ROW = 3
QTY_ROW = 4
OP_ROW = 2


def _exclusive_end(date_end: str) -> str:
    parsed = datetime.strptime(date_end.replace("-", ""), "%Y%m%d")
    return (parsed + timedelta(days=1)).strftime("%Y%m%d")


def _normalize_product(raw) -> str | None:
    if raw is None:
        return None
    code = str(raw).strip().replace(".0", "")
    if not code or code.lower().startswith("código"):
        return None
    return code


def _normalize_qty(raw) -> float:
    if raw is None:
        return 0.0
    try:
        return float(raw)
    except (TypeError, ValueError):
        return 0.0


@dataclass
class PlanilhaLot:
    month: str
    col: str
    op: str
    product_code: str
    qty: float


@dataclass
class ProductCompare:
    product_code: str
    planilha_un: float = 0.0
    ct70_un: float = 0.0
    months_planilha: set[str] = field(default_factory=set)
    lot_count: int = 0
    tipo: str = ""

    @property
    def delta_un(self) -> float:
        return self.planilha_un - self.ct70_un

    @property
    def delta_pct(self) -> float | None:
        if self.planilha_un == 0:
            return None
        return (self.delta_un / self.planilha_un) * 100.0

    @property
    def has_ct70(self) -> bool:
        return self.ct70_un > 0


def parse_planilha(path: Path) -> tuple[list[PlanilhaLot], dict[str, float]]:
    lots: list[PlanilhaLot] = []
    by_product: dict[str, float] = defaultdict(float)

    wb = openpyxl.load_workbook(
        path,
        read_only=True,
        data_only=True,
        keep_links=False,
    )
    month_names = [name for name, _, _ in MONTHS if name in wb.sheetnames]

    for sheet_name in month_names:
        ws = wb[sheet_name]
        product_row: tuple = ()
        qty_row: tuple = ()
        op_row: tuple = ()

        for row_idx, row in enumerate(
            ws.iter_rows(min_row=1, max_row=7, values_only=True),
            start=1,
        ):
            if row_idx == PRODUCT_ROW:
                product_row = row
            elif row_idx == QTY_ROW:
                qty_row = row
            elif row_idx == OP_ROW:
                op_row = row

        max_len = max(len(product_row), len(qty_row), len(op_row))
        for col_idx in range(LOT_COL_START - 1, max_len):
            product = _normalize_product(
                product_row[col_idx] if col_idx < len(product_row) else None
            )
            qty = _normalize_qty(qty_row[col_idx] if col_idx < len(qty_row) else None)
            if not product or qty <= 0:
                continue

            op_raw = op_row[col_idx] if col_idx < len(op_row) else None
            op = str(op_raw).strip().replace(".0", "") if op_raw else ""
            col_letter = openpyxl.utils.get_column_letter(col_idx + 1)

            lots.append(
                PlanilhaLot(
                    month=sheet_name,
                    col=col_letter,
                    op=op,
                    product_code=product,
                    qty=qty,
                )
            )
            by_product[product] += qty

    wb.close()
    return lots, dict(by_product)


SQL_CT70_BY_PRODUCT = """
WITH ct_inspecao_final AS (
    SELECT HB.HB_FILIAL, HB.HB_COD AS ct_inspecao
    FROM SHB010 HB
    WHERE HB.D_E_L_E_T_ = ' '
      AND UPPER(HB.HB_NOME) LIKE '%INSPE%FINAL%'
      AND HB.HB_FILIAL = '01'
),
apont_inspecao AS (
    SELECT
        LTRIM(RTRIM(SH6.H6_PRODUTO)) AS product_code,
        LTRIM(RTRIM(SB1.B1_TIPO)) AS product_type,
        SUM(CAST(SH6.H6_QTDPROD AS FLOAT)) AS qtd_milheiro
    FROM SH6010 SH6
    INNER JOIN SB1010 SB1
        ON SB1.B1_COD = SH6.H6_PRODUTO
       AND SB1.D_E_L_E_T_ = ' '
       AND SB1.B1_TIPO IN ('PA', 'PI')
    INNER JOIN SH1010 SH1
        ON SH1.H1_FILIAL = SH6.H6_FILIAL
       AND SH1.H1_CODIGO = SH6.H6_RECURSO
       AND SH1.D_E_L_E_T_ = ' '
    INNER JOIN ct_inspecao_final CIF
        ON CIF.HB_FILIAL = SH6.H6_FILIAL
       AND CIF.ct_inspecao = SH1.H1_CTRAB
    WHERE SH6.D_E_L_E_T_ = ' '
      AND SH6.H6_FILIAL = '01'
      AND SH6.H6_TIPO = 'P'
      AND SH6.H6_OP <> ''
      AND SH6.H6_PRODUTO <> ''
      AND SH6.H6_RECURSO <> ''
      AND SH6.H6_DTAPONT >= ?
      AND SH6.H6_DTAPONT < ?
    GROUP BY SH6.H6_PRODUTO, SB1.B1_TIPO
)
SELECT product_code, product_type, qtd_milheiro * 1000 AS total_un
FROM apont_inspecao
"""

SQL_CT70_BY_PRODUCT_MONTH = """
WITH ct_inspecao_final AS (
    SELECT HB.HB_FILIAL, HB.HB_COD AS ct_inspecao
    FROM SHB010 HB
    WHERE HB.D_E_L_E_T_ = ' '
      AND UPPER(HB.HB_NOME) LIKE '%INSPE%FINAL%'
      AND HB.HB_FILIAL = '01'
)
SELECT
    LTRIM(RTRIM(SH6.H6_PRODUTO)) AS product_code,
    SUM(CAST(SH6.H6_QTDPROD AS FLOAT)) * 1000 AS total_un
FROM SH6010 SH6
INNER JOIN SB1010 SB1
    ON SB1.B1_COD = SH6.H6_PRODUTO
   AND SB1.D_E_L_E_T_ = ' '
   AND SB1.B1_TIPO IN ('PA', 'PI')
INNER JOIN SH1010 SH1
    ON SH1.H1_FILIAL = SH6.H6_FILIAL
   AND SH1.H1_CODIGO = SH6.H6_RECURSO
   AND SH1.D_E_L_E_T_ = ' '
INNER JOIN ct_inspecao_final CIF
    ON CIF.HB_FILIAL = SH6.H6_FILIAL
   AND CIF.ct_inspecao = SH1.H1_CTRAB
WHERE SH6.D_E_L_E_T_ = ' '
  AND SH6.H6_FILIAL = '01'
  AND SH6.H6_TIPO = 'P'
  AND SH6.H6_OP <> ''
  AND SH6.H6_PRODUTO <> ''
  AND SH6.H6_RECURSO <> ''
  AND SH6.H6_PRODUTO = ?
  AND SH6.H6_DTAPONT >= ?
  AND SH6.H6_DTAPONT < ?
GROUP BY SH6.H6_PRODUTO
"""

SQL_CT70_BY_OP = """
WITH ct_inspecao_final AS (
    SELECT HB.HB_FILIAL, HB.HB_COD AS ct_inspecao
    FROM SHB010 HB
    WHERE HB.D_E_L_E_T_ = ' '
      AND UPPER(HB.HB_NOME) LIKE '%INSPE%FINAL%'
      AND HB.HB_FILIAL = '01'
)
SELECT
    LTRIM(RTRIM(SH6.H6_OP)) AS op,
    LTRIM(RTRIM(SH6.H6_PRODUTO)) AS product_code,
    SUM(CAST(SH6.H6_QTDPROD AS FLOAT)) * 1000 AS total_un
FROM SH6010 SH6
INNER JOIN SH1010 SH1
    ON SH1.H1_FILIAL = SH6.H6_FILIAL
   AND SH1.H1_CODIGO = SH6.H6_RECURSO
   AND SH1.D_E_L_E_T_ = ' '
INNER JOIN ct_inspecao_final CIF
    ON CIF.HB_FILIAL = SH6.H6_FILIAL
   AND CIF.ct_inspecao = SH1.H1_CTRAB
WHERE SH6.D_E_L_E_T_ = ' '
  AND SH6.H6_FILIAL = '01'
  AND SH6.H6_TIPO = 'P'
  AND SH6.H6_OP = ?
  AND SH6.H6_DTAPONT >= ?
  AND SH6.H6_DTAPONT < ?
GROUP BY SH6.H6_OP, SH6.H6_PRODUTO
"""


def fetch_ct70_products(date_start: str, date_end: str) -> dict[str, tuple[float, str]]:
    ini = date_start.replace("-", "")
    fim_ex = _exclusive_end(date_end)
    with PpmQueryRepository() as repo:
        rows = repo.execute_query(SQL_CT70_BY_PRODUCT, (ini, fim_ex)) or []
    return {
        str(r.get("product_code") or "").strip(): (
            float(r.get("total_un") or 0),
            str(r.get("product_type") or "").strip(),
        )
        for r in rows
    }


def main() -> None:
    if not SHEET_PATH.exists():
        raise SystemExit(f"Planilha não encontrada: {SHEET_PATH}")

    lots, planilha_by_product = parse_planilha(SHEET_PATH)
    ct70 = fetch_ct70_products("2026-01-01", "2026-05-31")

    compares: dict[str, ProductCompare] = {}
    for code, qty in planilha_by_product.items():
        ct_qty, tipo = ct70.get(code, (0.0, ""))
        compares[code] = ProductCompare(
            product_code=code,
            planilha_un=qty,
            ct70_un=ct_qty,
            tipo=tipo,
            lot_count=sum(1 for lot in lots if lot.product_code == code),
            months_planilha={lot.month for lot in lots if lot.product_code == code},
        )

    # Produtos só no CT-70 (não na planilha)
    only_ct70 = {
        code: data
        for code, data in ct70.items()
        if code not in planilha_by_product and data[0] > 0
    }

    sorted_compares = sorted(compares.values(), key=lambda x: -x.planilha_un)

    print("=" * 90)
    print("PLANILHA RQ 005 × CT-70 (filial 01, jan–mai/2026)")
    print("Planilha: soma linha 4 (Quantidade do Lote) | CT-70: SUM(H6_QTDPROD) CT inspeção final")
    print("=" * 90)

    total_plan = sum(c.planilha_un for c in compares.values())
    total_ct = sum(c.ct70_un for c in compares.values())
    print(f"\nTotais (produtos presentes na planilha): planilha={total_plan:,.0f} un. | CT-70={total_ct:,.0f} un.")

    # Categorias
    zero_ct = [c for c in sorted_compares if not c.has_ct70]
    partial = [
        c for c in sorted_compares
        if c.has_ct70 and c.delta_pct is not None and abs(c.delta_pct) > 5
    ]
    aligned = [
        c for c in sorted_compares
        if c.has_ct70 and c.delta_pct is not None and abs(c.delta_pct) <= 5
    ]

    print(f"\n--- Produtos na planilha SEM apontamento no CT-70: {len(zero_ct)} ---")
    print(f"{'Código':<12} {'Tipo':<4} {'Planilha':>12} {'Lotes':>6} {'Meses'}")
    print("-" * 70)
    for c in zero_ct[:30]:
        meses = ",".join(sorted(c.months_planilha))
        print(f"{c.product_code:<12} {'?':<4} {c.planilha_un:>12,.0f} {c.lot_count:>6} {meses}")
    if len(zero_ct) > 30:
        print(f"... +{len(zero_ct) - 30} produtos")
    print(f"Subtotal planilha (sem CT-70): {sum(c.planilha_un for c in zero_ct):,.0f} un.")

    print(f"\n--- PIs críticos (50232xxx / 50233xxx) ---")
    pi_codes = ["50232465", "50233615", "50233616", "550232465"]
    for code in pi_codes:
        c = compares.get(code)
        if c:
            print(
                f"  {code}: planilha={c.planilha_un:,.0f} | CT-70={c.ct70_un:,.0f} | "
                f"lotes={c.lot_count} | meses={','.join(sorted(c.months_planilha))}"
            )
        elif code in ct70:
            print(f"  {code}: planilha=0 | CT-70={ct70[code][0]:,.0f}")
        else:
            print(f"  {code}: planilha=0 | CT-70=0")

    print(f"\n--- Top 20 desvios (planilha >> CT-70) ---")
    print(f"{'Código':<12} {'Tipo':<4} {'Planilha':>12} {'CT-70':>12} {'Δ%':>8} {'Lotes':>6}")
    print("-" * 60)
    for c in sorted(partial, key=lambda x: -x.delta_un)[:20]:
        pct = f"{c.delta_pct:+.0f}%" if c.delta_pct is not None else "n/a"
        print(
            f"{c.product_code:<12} {c.tipo or '?':<4} {c.planilha_un:>12,.0f} "
            f"{c.ct70_un:>12,.0f} {pct:>8} {c.lot_count:>6}"
        )

    print(f"\n--- Alinhados (|Δ| ≤ 5%): {len(aligned)} produtos ---")
    print(f"Subtotal: planilha={sum(c.planilha_un for c in aligned):,.0f} | CT-70={sum(c.ct70_un for c in aligned):,.0f}")

    print(f"\n--- Produtos apontados no CT-70 mas AUSENTES na planilha: {len(only_ct70)} ---")
    top_only_ct = sorted(only_ct70.items(), key=lambda x: -x[1][0])[:15]
    for code, (qty, tipo) in top_only_ct:
        print(f"  {code} ({tipo}): {qty:,.0f} un.")

    # Verificação por OP para PIs zero CT
    print("\n--- Amostra: OPs dos PIs na planilha vs apontamento CT-70 no mês ---")
    with PpmQueryRepository() as repo:
        for code in ["50232465", "50233615", "50233616", "550232465"]:
            pi_lots = [lot for lot in lots if lot.product_code == code][:3]
            for lot in pi_lots:
                month_cfg = next(m for m in MONTHS if m[0] == lot.month)
                ini = month_cfg[1].replace("-", "")
                fim_ex = _exclusive_end(month_cfg[2])
                op_rows = repo.execute_query(
                    SQL_CT70_BY_OP,
                    (lot.op, ini, fim_ex),
                ) or []
                ct_op = sum(float(r.get("total_un") or 0) for r in op_rows)
                ct_prod = repo.execute_one(
                    SQL_CT70_BY_PRODUCT_MONTH,
                    (code, ini, fim_ex),
                ) or {}
                print(
                    f"  {lot.month} OP {lot.op} prod {code} col {lot.col}: "
                    f"planilha={lot.qty:,.0f} | CT OP={ct_op:,.0f} | CT prod mês={float(ct_prod.get('total_un') or 0):,.0f}"
                )

    # Resumo mensal PI
    print("\n--- PIs na planilha por mês vs CT-70 ---")
    print(f"{'Mês':<6} {'Plan PI':>12} {'CT PI':>12} {'Plan total':>14} {'CT total':>12}")
    print("-" * 60)
    with PpmQueryRepository() as repo:
        for sheet, start, end in MONTHS:
            ini = start.replace("-", "")
            fim_ex = _exclusive_end(end)
            month_lots = [lot for lot in lots if lot.month == sheet]
            plan_total = sum(lot.qty for lot in month_lots)
            plan_pi = sum(
                lot.qty
                for lot in month_lots
                if lot.product_code.startswith(("50232", "50233", "55023"))
            )
            rows = repo.execute_query(SQL_CT70_BY_PRODUCT, (ini, fim_ex)) or []
            ct_total = sum(float(r.get("total_un") or 0) for r in rows)
            ct_pi = sum(
                float(r.get("total_un") or 0)
                for r in rows
                if str(r.get("product_code") or "").startswith(("50232", "50233"))
            )
            print(
                f"{sheet:<6} {plan_pi:>12,.0f} {ct_pi:>12,.0f} "
                f"{plan_total:>14,.0f} {ct_total:>12,.0f}"
            )


if __name__ == "__main__":
    main()
