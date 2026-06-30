from __future__ import annotations

import io
import zipfile
from pathlib import Path

import pytest
from openpyxl import load_workbook

from app.domain.services.quality_action_plans.rnc_8d_excel_export_service import (
    build_rnc_8d_workbook,
)
from app.domain.services.quality_action_plans.rnc_8d_ooxml_merge_service import (
    merge_template_visual_assets,
)

_TEMPLATE = (
    Path(__file__).resolve().parents[1]
    / "app"
    / "content"
    / "templates"
    / "quality"
    / "weg_wfr20997_template.xlsx"
)


@pytest.mark.skipif(not _TEMPLATE.is_file(), reason="Template WEG ausente")
def test_merge_template_visual_assets_preserves_template_ooxml_shell():
    from openpyxl import Workbook

    buffer = io.BytesIO()
    workbook = Workbook()
    workbook.active.title = "R8D"
    workbook.create_sheet("Anexos(Evidencias)")
    workbook.save(buffer)

    merged = merge_template_visual_assets(
        template_path=_TEMPLATE,
        filled_bytes=buffer.getvalue(),
    )
    with zipfile.ZipFile(io.BytesIO(merged)) as merged_zip:
        names = set(merged_zip.namelist())
        assert merged_zip.namelist()[0] == "[Content_Types].xml"
        assert "xl/drawings/drawing1.xml" in names
        assert "xl/media/image1.png" in names
        assert "xl/sharedStrings.xml" not in names
        assert "xl/calcChain.xml" not in names

        content_types = merged_zip.read("[Content_Types].xml").decode()
        assert "ns0:" not in content_types
        assert "/xl/sharedStrings.xml" not in content_types

        workbook_rels = merged_zip.read("xl/_rels/workbook.xml.rels").decode()
        assert "sharedStrings.xml" not in workbook_rels
        assert "calcChain.xml" not in workbook_rels
        assert 'Target="worksheets/sheet1.xml"' in workbook_rels

        sheet1 = merged_zip.read("xl/worksheets/sheet1.xml").decode()
        assert 'drawing r:id="rId2"' in sheet1
        assert "ignoredErrors" in sheet1


@pytest.mark.skipif(not _TEMPLATE.is_file(), reason="Template WEG ausente")
def test_build_rnc_8d_workbook_export_is_excel_compatible():
    detail = {
        "plan": {
            "branch_code": "01",
            "client_nc_registry": "217383667",
            "template_payload": {"report_date": "2028-06-29"},
        },
        "team_members": [],
        "actions": [],
        "five_whys": {
            "occurrence_whys": [
                {"question": "Por quê ocorreu?", "answer": "Falha na montagem."},
            ],
        },
    }
    content = build_rnc_8d_workbook(detail, template_key="weg_wfr20997")

    with zipfile.ZipFile(io.BytesIO(content)) as exported:
        names = set(exported.namelist())
        assert "xl/drawings/drawing1.xml" in names
        assert "xl/sharedStrings.xml" not in names
        content_types = exported.read("[Content_Types].xml").decode()
        assert "ns0:" not in content_types

        sheet1 = exported.read("xl/worksheets/sheet1.xml").decode()
        assert 'drawing r:id="rId2"' in sheet1
        assert "IF(C1=1" in sheet1

    workbook = load_workbook(io.BytesIO(content), data_only=True)
    ws = workbook["R8D"]
    assert ws["I4"].value == "217383667"
    assert ws["E44"].value == "Falha na montagem."
