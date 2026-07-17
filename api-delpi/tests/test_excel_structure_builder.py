"""Excel da estrutura — intermediários sem MP amarrada não podem sumir da planilha."""

from openpyxl import load_workbook

from app.application.services.excel_structure_builder import ExcelStructureBuilder
from app.domain.entities.product.bom_node import BomNode


def _node(code: str, description: str, node_type: str, unit: str, quantity: float = 1.0, components=None) -> BomNode:
    return BomNode(
        code=code,
        description=description,
        type=node_type,
        unit=unit,
        quantity=quantity,
        components=list(components or []),
    )


def _build_tree() -> BomNode:
    return _node(
        "90261699",
        "CHICOTE DE LIGACAO",
        "PA",
        "MI",
        components=[
            _node("50210723", "CA0,75PRET-00425/14/10-0000-0000", "PI", "MI"),
            _node("50210724", "CA0,75BRAN-00425/14/10-0000-0000", "PI", "MI"),
            _node(
                "50221531",
                "CA0,75MRBN-00255/14/06-0000-0200",
                "PI",
                "MI",
                components=[
                    _node("10020008", "CABO PVC 105C 0,75MM2", "MP", "MT", 255),
                    _node("10080002", "TERM. BANDEIRA 6,30X0,08", "MP", "PC", 1000),
                ],
            ),
        ],
    )


def _sheet_component_codes(stream) -> list[str]:
    workbook = load_workbook(stream)
    sheet = workbook.active

    return [
        str(row[4] or "").strip()
        for row in sheet.iter_rows(min_row=2, values_only=True)
        if str(row[4] or "").strip()
    ]


def test_intermediates_without_raw_material_appear_in_excel():
    stream = ExcelStructureBuilder.build(_build_tree())
    codes = _sheet_component_codes(stream)

    assert "50210723" in codes
    assert "50210724" in codes
    assert "10020008" in codes
    assert "10080002" in codes


def test_intermediate_with_children_still_expands_to_raw_materials():
    stream = ExcelStructureBuilder.build(_build_tree())
    codes = _sheet_component_codes(stream)

    # PI com filhos continua explodido em MPs, não listado como linha própria.
    assert "50221531" not in codes
