from __future__ import annotations

import io
import re
import zipfile
from pathlib import Path

import pytest
from openpyxl import load_workbook

from app.domain.services.quality_action_plans.rnc_8d_excel_export_service import (
    build_rnc_8d_workbook,
)
from app.domain.services.quality_action_plans.rnc_8d_template_fill_service import (
    fill_template_workbook,
)

_TEMPLATE = (
    Path(__file__).resolve().parents[1]
    / "app"
    / "content"
    / "templates"
    / "quality"
    / "weg_wfr20997_template.xlsx"
)


@pytest.mark.skipif(not _TEMPLATE.is_file(), reason="Template WEG ausente")
def test_fill_template_workbook_preserves_template_structure():
    with zipfile.ZipFile(_TEMPLATE, "r") as template_zip:
        template_sheet = template_zip.read("xl/worksheets/sheet1.xml").decode()
        template_cell_count = len(re.findall(r'<c r="', template_sheet))

    filled = fill_template_workbook(
        template_path=_TEMPLATE,
        cell_values={"I4": "217383667", "D63": "Resolvido", "D70": "09/02/2026"},
    )

    with zipfile.ZipFile(io.BytesIO(filled)) as exported:
        assert exported.namelist()[0] == "[Content_Types].xml"
        assert "xl/drawings/drawing1.xml" in exported.namelist()
        assert "xl/sharedStrings.xml" in exported.namelist()
        assert "xl/calcChain.xml" not in exported.namelist()
        assert exported.getinfo("xl/media/image1.png").compress_type == zipfile.ZIP_STORED

        sheet1 = exported.read("xl/worksheets/sheet1.xml").decode()
        assert len(re.findall(r'<c r="', sheet1)) == template_cell_count
        assert 'drawing r:id="rId2"' in sheet1
        assert "IF(C1=1" in sheet1
        assert "217383667" in sheet1
        assert "Resolvido" in sheet1

        workbook = load_workbook(io.BytesIO(filled), data_only=True)
        ws = workbook["R8D"]
        assert ws["I4"].value == "217383667"
        assert ws["D63"].value == "Resolvido"
        assert ws["D70"].value == "09/02/2026"
        assert str(workbook["R8D"]["D71"].value or "").startswith("Como este assunto")


@pytest.mark.skipif(not _TEMPLATE.is_file(), reason="Template WEG ausente")
def test_build_rnc_8d_workbook_export_is_excel_compatible():
    detail = {
        "plan": {
            "branch_code": "01",
            "client_nc_registry": "217383667",
            "effectiveness_notes": "Problema resolvido com retrabalho.",
            "template_payload": {
                "report_date": "2028-06-29",
                "effectiveness": {
                    "ok_material_date": "2026-02-09",
                    "new_parts_identification": "Etiqueta verde",
                    "verification_responsible": "Carla Demeneck",
                    "verification_date": "2026-02-12",
                },
                "preventive": {
                    "how_avoid_future": "Revisar desvio no PCP.",
                    "other_processes_products": "Aplicar em todos os chicotes.",
                    "evaluation_responsible": "Elaine Marquardt",
                    "evaluation_completion_date": "2026-04-05",
                },
                "documentation_updates": [
                    {"document": "IT-123", "responsible": "Laercio Koch", "date": "2026-03-01"},
                ],
                "client_closure_note": "NC encerrada após validação do cliente.",
            },
        },
        "team_members": [],
        "actions": [],
        "five_whys": {
            "occurrence_whys": [
                {"question": "Por quê ocorreu?", "answer": "Falha na montagem."},
            ],
        },
    }
    content = build_rnc_8d_workbook(detail, template_key="weg_wfr20997")

    with zipfile.ZipFile(io.BytesIO(content)) as exported:
        assert "xl/drawings/drawing1.xml" in set(exported.namelist())
        sheet1 = exported.read("xl/worksheets/sheet1.xml").decode()
        assert 'drawing r:id="rId2"' in sheet1

    workbook = load_workbook(io.BytesIO(content), data_only=True)
    ws = workbook["R8D"]
    assert ws["I4"].value == "217383667"
    assert ws["E44"].value == "Falha na montagem."
    assert ws["D63"].value == "Problema resolvido com retrabalho."
    assert ws["D72"].value == "Revisar desvio no PCP."
    assert ws["F85"].value == "IT-123"
    assert ws["D107"].value == "NC encerrada após validação do cliente."
