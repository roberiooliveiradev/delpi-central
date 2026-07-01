from __future__ import annotations

import io
import zipfile
from pathlib import Path

import pytest

try:
    from PIL import Image
except ImportError:
    Image = None  # type: ignore[assignment,misc]

from app.domain.services.quality_action_plans.rnc_8d_excel_export_service import (
    build_rnc_8d_workbook,
    is_image_evidence,
    resolve_rnc_8d_template_path,
)

_TEMPLATE_PATH = resolve_rnc_8d_template_path("weg_wfr20997")


def _minimal_png_bytes() -> bytes:
    if Image is None:
        pytest.skip("Pillow ausente")
    buffer = io.BytesIO()
    Image.new("RGB", (8, 8), color=(200, 40, 40)).save(buffer, format="PNG")
    return buffer.getvalue()


def test_is_image_evidence_detects_mime_and_type():
    assert is_image_evidence({"mime_type": "image/png", "type": "other"})
    assert is_image_evidence({"type": "image"})
    assert not is_image_evidence({"mime_type": "application/pdf", "type": "pdf"})


@pytest.mark.skipif(not _TEMPLATE_PATH.is_file(), reason="Template 8D ausente")
def test_build_rnc_8d_workbook_fills_registry_cell():
    detail = {
        "plan": {
            "client_nc_registry": "215571003",
            "customer_name": "Cliente industrial",
            "product_code": "14297268",
            "product_description": "CHICOTE 3 SILICONE 150mm",
            "batch_number": "90262776",
            "branch_code": "01",
            "template_payload": {
                "purchase_order": "5500044658 / 09770",
                "invoice_number": "000092387-1",
                "client_batch": "10019632175",
                "disposition": "Rejeitado",
            },
        },
        "five_whys": {
            "occurrence_whys": ["Falha na impressão"],
        },
        "team_members": [
            {"member_name": "Rodrigo J. Cozer", "department": "Qualidade", "is_leader": True},
        ],
        "actions": [],
    }

    content = build_rnc_8d_workbook(detail)
    assert isinstance(content, bytes)
    assert len(content) > 10_000

    from openpyxl import load_workbook
    import io

    wb = load_workbook(io.BytesIO(content), data_only=True)
    ws = wb["R8D"]
    assert ws["I4"].value == "215571003"
    assert "14297268" in str(ws["E6"].value)
    assert ws["J8"].value == "10019632175"


@pytest.mark.skipif(not _TEMPLATE_PATH.is_file(), reason="Template 8D ausente")
def test_build_rnc_8d_workbook_preserves_weg_drawings():
    detail = {
        "plan": {"client_nc_registry": "215571003", "branch_code": "01", "template_payload": {}},
        "team_members": [],
        "actions": [],
    }
    content = build_rnc_8d_workbook(detail, template_key="weg_wfr20997")
    with zipfile.ZipFile(io.BytesIO(content)) as exported:
        names = set(exported.namelist())
        assert "xl/drawings/drawing1.xml" in names
        sheet1 = exported.read("xl/worksheets/sheet1.xml").decode()
        assert 'drawing r:id="rId2"' in sheet1


@pytest.mark.skipif(not _TEMPLATE_PATH.is_file(), reason="Template 8D ausente")
def test_build_rnc_8d_workbook_writes_why_answers_only():
    detail = {
        "plan": {"branch_code": "01", "template_payload": {}},
        "team_members": [],
        "actions": [],
        "five_whys": {
            "occurrence_whys": [
                {"question": "Por quê ocorreu?", "answer": "Falha na montagem."},
            ],
        },
    }
    content = build_rnc_8d_workbook(detail, template_key="weg_wfr20997")
    from openpyxl import load_workbook

    ws = load_workbook(io.BytesIO(content), data_only=True)["R8D"]
    assert ws["E44"].value == "Falha na montagem."


@pytest.mark.skipif(not _TEMPLATE_PATH.is_file(), reason="Template 8D ausente")
def test_build_rnc_8d_workbook_maps_customer_and_delpi_contacts_to_weg_cells():
    detail = {
        "plan": {
            "branch_code": "01",
            "customer_contact": "Igor Sfalsin Zamperlini",
            "customer_contact_email": "wmo-rnc@weg.net",
            "delpi_contact_name": "Laercio Koch",
            "template_payload": {"contact_phone": "47 3370 5502"},
        },
        "team_members": [],
        "actions": [],
    }
    content = build_rnc_8d_workbook(detail, template_key="weg_wfr20997")
    from openpyxl import load_workbook

    ws = load_workbook(io.BytesIO(content), data_only=True)["R8D"]
    assert ws["J5"].value == "Laercio Koch"
    assert ws["J6"].value == "47 3370 5502"
    assert ws["G21"].value == "Igor Sfalsin Zamperlini"
    assert ws["J21"].value == "wmo-rnc@weg.net"


@pytest.mark.skipif(not _TEMPLATE_PATH.is_file(), reason="Template 8D ausente")
def test_build_rnc_8d_workbook_formats_dates_as_brazilian_text():
    detail = {
        "plan": {
            "branch_code": "01",
            "template_payload": {"report_date": "2028-06-29"},
        },
        "team_members": [],
        "actions": [],
    }
    content = build_rnc_8d_workbook(detail, template_key="weg_wfr20997")
    from openpyxl import load_workbook

    ws = load_workbook(io.BytesIO(content), data_only=True)["R8D"]
    assert ws["K1"].value == "29/06/2028"


@pytest.mark.skipif(not _TEMPLATE_PATH.is_file(), reason="Template 8D ausente")
def test_build_rnc_8d_workbook_maps_lower_sections_to_content_cells():
    detail = {
        "plan": {
            "branch_code": "01",
            "effectiveness_notes": "Problema resolvido com retrabalho.",
            "template_payload": {
                "effectiveness": {
                    "ok_material_date": "2026-02-09",
                    "new_parts_identification": "Etiqueta verde",
                    "verification_responsible": "Carla Demeneck",
                    "verification_date": "2026-02-12",
                },
                "preventive": {
                    "how_avoid_future": "Revisar desvio no PCP.",
                    "other_processes_products": "Aplicar em todos os chicotes.",
                    "evaluation_responsible": "Elaine Marquardt",
                    "evaluation_completion_date": "2026-04-05",
                },
                "documentation_updates": [
                    {"document": "IT-123", "responsible": "Laercio Koch", "date": "2026-03-01"},
                ],
                "client_closure_note": "NC encerrada após validação do cliente.",
            },
        },
        "team_members": [],
        "actions": [],
    }
    content = build_rnc_8d_workbook(detail, template_key="weg_wfr20997")
    from openpyxl import load_workbook

    ws_values = load_workbook(io.BytesIO(content), data_only=True)["R8D"]
    ws_formulas = load_workbook(io.BytesIO(content))["R8D"]

    assert ws_values["D63"].value == "Problema resolvido com retrabalho."
    assert ws_values["D70"].value == "09/02/2026"
    assert ws_values["F70"].value == "Etiqueta verde"
    assert ws_values["J70"].value == "Carla Demeneck"
    assert ws_values["L70"].value == "12/02/2026"
    assert ws_values["D72"].value == "Revisar desvio no PCP."
    assert ws_values["D77"].value == "Aplicar em todos os chicotes."
    assert ws_values["D83"].value == "Elaine Marquardt"
    assert ws_values["I83"].value == "05/04/2026"
    assert ws_values["F85"].value == "IT-123"
    assert ws_values["I85"].value == "Laercio Koch"
    assert ws_values["K85"].value == "01/03/2026"
    assert ws_values["D107"].value == "NC encerrada após validação do cliente."
    assert str(ws_formulas["D71"].value or "").startswith("=IF(C1=1")
    assert str(ws_formulas["D84"].value or "").startswith("=IF(C1=1")


@pytest.mark.skipif(not _TEMPLATE_PATH.is_file(), reason="Template 8D ausente")
def test_build_rnc_8d_workbook_embeds_annex_images():
    detail = {
        "plan": {"client_nc_registry": "215571003", "branch_code": "01", "template_payload": {}},
        "team_members": [],
        "actions": [],
    }
    content = build_rnc_8d_workbook(
        detail,
        image_annexes=[
            {
                "file_name": "foto-nc.png",
                "description": "Evidência da não conformidade",
                "content": _minimal_png_bytes(),
            }
        ],
    )
    from openpyxl import load_workbook

    wb = load_workbook(io.BytesIO(content))
    assert "Anexos(Evidencias)" in wb.sheetnames
    annex = wb["Anexos(Evidencias)"]
    assert annex["D3"].value == "foto-nc.png"
    assert len(annex._images) >= 1
