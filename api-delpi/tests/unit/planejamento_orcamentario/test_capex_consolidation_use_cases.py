"""Testes Fase 2D.1 — consolidação gerencial CAPEX e exportação Excel."""
from __future__ import annotations

from copy import deepcopy
from datetime import date
from decimal import Decimal
from io import BytesIO
from typing import Any
from uuid import uuid4

import pytest
from openpyxl import load_workbook

from app.application.use_cases.planejamento_orcamentario.budget_planning_use_cases import (
    BudgetActor,
)
from app.application.use_cases.planejamento_orcamentario.capex_consolidation_use_cases import (
    CapexConsolidationUseCases,
)
from app.domain.services.planejamento_orcamentario.exceptions import (
    BudgetUserNotAuthorizedError,
    CapexConsolidationCurrencyConflictError,
    CapexConsolidationForbiddenError,
    CapexExportForbiddenError,
    CapexInvestmentCostCenterForbiddenError,
)
from tests.unit.planejamento_orcamentario.fake_po_repository_helpers import (
    cc_lookup_for_seed,
    default_po_cc_seed_with_es,
    filter_unit_cost_center_pairs,
)


class FakeRepo:
    def __init__(self) -> None:
        self.exercises: dict[str, dict] = {}
        self.ccs: dict[tuple[str, str], dict] = default_po_cc_seed_with_es()
        self.units = {
            "01": {"code": "01", "name": "SC"},
            "02": {"code": "02", "name": "ES"},
        }
        self.areas = {
            "PROD": {"code": "PROD", "name": "Produção"},
            "ENG": {"code": "ENG", "name": "Engenharia"},
        }
        self.categories: dict[str, dict] = {}
        self.responsibilities: list[dict] = []
        self.investments: dict[str, dict] = {}
        self.plans: dict[str, dict] = {}
        self.audits: list[dict] = []

    def seed_exercise(self, *, year: int = 2027) -> str:
        eid = str(uuid4())
        self.exercises[eid] = {
            "id": eid,
            "year": year,
            "name": f"PO {year}",
            "status": "open",
            "is_active": True,
        }
        return eid

    def seed_category(self, *, code: str = "FERR", name: str = "Ferramentas") -> str:
        cid = str(uuid4())
        self.categories[cid] = {
            "id": cid,
            "code": code,
            "name": name,
            "is_active": True,
        }
        return cid

    def seed_responsibility(self, *, user_sub: str, exercise_id: str, cost_center_id: str):
        cc = cc_lookup_for_seed(self.ccs, cost_center_id)
        self.responsibilities.append(
            {
                "id": str(uuid4()),
                "user_sub": user_sub,
                "exercise_id": exercise_id,
                "module": "capex",
                "cost_center_id": cost_center_id,
                "unit_id": cc["unit_code"],
                "area_id": cc["area_code"],
                "is_active": True,
            }
        )

    def seed_investment(
        self,
        *,
        exercise_id: str,
        cost_center_id: str = "205",
        category_id: str | None = None,
        amount: str = "1000.00",
        priority: str = "2",
        origin: str = "national",
        required_date: str = "2027-06-01",
        status: str = "draft",
        currency: str = "BRL",
        description: str = "Investimento",
        incomplete: bool = False,
    ) -> str:
        iid = str(uuid4())
        cc = cc_lookup_for_seed(self.ccs, cost_center_id)
        row = {
            "id": iid,
            "exercise_id": exercise_id,
            "unit_id": cc["unit_code"],
            "area_id": cc["area_code"],
            "cost_center_id": cost_center_id,
            "category_id": None if incomplete else category_id,
            "description": description,
            "estimated_amount": amount,
            "currency": currency,
            "required_date": required_date,
            "priority": priority,
            "origin": origin,
            "probable_supplier_name": "Fornecedor X",
            "status": status,
            "review_status": "pending",
            "version": 1,
            "updated_at": "2026-08-01T10:00:00Z",
            "created_at": "2026-07-01T10:00:00Z",
        }
        if incomplete:
            row["category_id"] = None
        self.investments[iid] = row
        return iid

    def seed_plan(
        self,
        *,
        exercise_id: str,
        cost_center_id: str,
        status: str = "draft",
        submitted_by: str | None = None,
    ) -> str:
        pid = str(uuid4())
        cc = cc_lookup_for_seed(self.ccs, cost_center_id)
        self.plans[pid] = {
            "id": pid,
            "exercise_id": exercise_id,
            "unit_id": cc["unit_code"],
            "area_id": cc["area_code"],
            "cost_center_id": cost_center_id,
            "status": status,
            "submitted_by": submitted_by,
            "version": 1,
        }
        return pid

    def get_exercise(self, exercise_id: str):
        return deepcopy(self.exercises.get(exercise_id))

    def get_exercise_by_year(self, year: int):
        for ex in self.exercises.values():
            if int(ex["year"]) == int(year):
                return deepcopy(ex)
        return None

    def list_budget_responsibilities_for_user(self, **kwargs):
        items = []
        for r in self.responsibilities:
            if r["user_sub"] != kwargs["user_sub"]:
                continue
            if kwargs.get("module") and r["module"] != kwargs["module"]:
                continue
            if kwargs.get("active_only", True) and not r["is_active"]:
                continue
            items.append(deepcopy(r))
        return items

    def list_capex_consolidation_rows(self, **kwargs):
        items: list[dict[str, Any]] = []
        for inv in self.investments.values():
            if inv["exercise_id"] != kwargs["exercise_id"]:
                continue
            if inv.get("status") == "archived":
                continue
            if kwargs.get("unit_id") and inv["unit_id"] != kwargs["unit_id"]:
                continue
            if kwargs.get("area_id") and inv["area_id"] != kwargs["area_id"]:
                continue
            if kwargs.get("cost_center_id") and inv["cost_center_id"] != kwargs["cost_center_id"]:
                continue
            scoped = filter_unit_cost_center_pairs(
                [inv], kwargs.get("unit_cost_center_pairs")
            )
            if kwargs.get("unit_cost_center_pairs") is not None and not scoped:
                continue
            if kwargs.get("cost_center_ids") is not None and kwargs.get("unit_cost_center_pairs") is None:
                if not kwargs["cost_center_ids"]:
                    return [], 0
                if inv["cost_center_id"] not in kwargs["cost_center_ids"]:
                    continue
            if kwargs.get("category_id") and inv.get("category_id") != kwargs["category_id"]:
                continue
            if kwargs.get("priority") and inv.get("priority") != kwargs["priority"]:
                continue
            if kwargs.get("origin") and inv.get("origin") != kwargs["origin"]:
                continue
            plan = next(
                (
                    p
                    for p in self.plans.values()
                    if p["exercise_id"] == inv["exercise_id"]
                    and p["cost_center_id"] == inv["cost_center_id"]
                    and p.get("unit_id") == inv.get("unit_id")
                ),
                None,
            )
            plan_status = plan["status"] if plan else "draft"
            if kwargs.get("plan_status"):
                if kwargs["plan_status"] == "draft":
                    if plan_status != "draft":
                        continue
                elif plan_status != kwargs["plan_status"]:
                    continue
            req = inv.get("required_date")
            if kwargs.get("required_date_from") and (not req or str(req) < kwargs["required_date_from"]):
                continue
            if kwargs.get("required_date_to") and (not req or str(req) > kwargs["required_date_to"]):
                continue

            cat = self.categories.get(inv.get("category_id") or "")
            row = deepcopy(inv)
            row["plan_id"] = plan["id"] if plan else None
            row["plan_status"] = plan_status
            row["plan_submitted_by"] = plan.get("submitted_by") if plan else None
            row["category_code"] = cat["code"] if cat else None
            row["category_name"] = cat["name"] if cat else None
            row["unit_name"] = self.units.get(inv["unit_id"], {}).get("name")
            row["area_name"] = self.areas.get(inv["area_id"], {}).get("name")
            try:
                row["cost_center_name"] = cc_lookup_for_seed(
                    self.ccs, inv["cost_center_id"]
                ).get("name")
            except KeyError:
                row["cost_center_name"] = None
            row["responsible"] = row.get("plan_submitted_by")
            items.append(row)

        total = len(items)
        sort_by = kwargs.get("sort_by", "updated_at")
        reverse = str(kwargs.get("sort_dir", "desc")).lower() != "asc"
        items.sort(key=lambda r: str(r.get(sort_by) or ""), reverse=reverse)
        if kwargs.get("limit") is not None:
            offset = int(kwargs.get("offset") or 0)
            limit = int(kwargs["limit"])
            items = items[offset : offset + limit]
        return items, total

    def append_audit(self, **kwargs):
        self.audits.append(kwargs)


def _admin() -> BudgetActor:
    return BudgetActor(
        user_id="admin",
        user_name="Admin",
        permissions=frozenset(
            {
                "planejamento-orcamentario.admin",
                "planejamento-orcamentario.access",
            }
        ),
    )


def _consolidator() -> BudgetActor:
    return BudgetActor(
        user_id="cons",
        user_name="Consolidador",
        permissions=frozenset(
            {
                "planejamento-orcamentario.access",
                "planejamento-orcamentario.capex.consolidation.view",
                "planejamento-orcamentario.capex.export",
            }
        ),
    )


def _manager(user_id: str = "mgr") -> BudgetActor:
    return BudgetActor(
        user_id=user_id,
        user_name="Gestor",
        permissions=frozenset(
            {
                "planejamento-orcamentario.access",
                "planejamento-orcamentario.capex.submit",
            }
        ),
    )


def _no_perm() -> BudgetActor:
    return BudgetActor(user_id="x", user_name="X", permissions=frozenset())


@pytest.fixture
def repo() -> FakeRepo:
    return FakeRepo()


@pytest.fixture
def uc(repo: FakeRepo) -> CapexConsolidationUseCases:
    return CapexConsolidationUseCases(repository=repo)  # type: ignore[arg-type]


def test_summary_empty(uc, repo):
    eid = repo.seed_exercise()
    result = uc.get_summary(_consolidator(), exercise_id=eid, audit=False)
    s = result["summary"]
    assert s["investment_count"] == 0
    assert s["total_estimated_amount"] == "0.00"
    assert s["cost_center_count"] == 0
    assert s["incomplete_investment_count"] == 0


def test_summary_multiple_cost_centers(uc, repo):
    eid = repo.seed_exercise()
    cat = repo.seed_category()
    repo.seed_investment(exercise_id=eid, cost_center_id="205", category_id=cat, amount="1000.00")
    repo.seed_investment(exercise_id=eid, cost_center_id="210", category_id=cat, amount="2500.50")
    repo.seed_plan(exercise_id=eid, cost_center_id="205", status="submitted")
    repo.seed_plan(exercise_id=eid, cost_center_id="210", status="approved")
    result = uc.get_summary(_admin(), exercise_id=eid, audit=True)
    s = result["summary"]
    assert s["investment_count"] == 2
    assert s["cost_center_count"] == 2
    assert Decimal(s["total_estimated_amount"]) == Decimal("3500.50")
    assert s["plans_submitted_count"] == 1
    assert s["plans_approved_count"] == 1
    assert Decimal(s["approved_amount"]) == Decimal("2500.50")
    assert Decimal(s["in_review_amount"]) == Decimal("1000.00")
    assert repo.audits and repo.audits[-1]["action"] == "capex_consolidation.summary_viewed"


def test_summary_exclui_investimento_reprovado_do_aprovado(uc, repo):
    eid = repo.seed_exercise()
    cat = repo.seed_category()
    approved = repo.seed_investment(
        exercise_id=eid, cost_center_id="205", category_id=cat, amount="800.00"
    )
    rejected = repo.seed_investment(
        exercise_id=eid, cost_center_id="205", category_id=cat, amount="200.00"
    )
    repo.investments[approved]["review_status"] = "approved"
    repo.investments[rejected]["review_status"] = "rejected"
    repo.seed_plan(exercise_id=eid, cost_center_id="205", status="approved")
    result = uc.get_summary(_admin(), exercise_id=eid, audit=False)
    s = result["summary"]
    assert Decimal(s["total_estimated_amount"]) == Decimal("1000.00")
    assert Decimal(s["approved_amount"]) == Decimal("800.00")


def test_groupings(uc, repo):
    eid = repo.seed_exercise()
    cat_a = repo.seed_category(code="A", name="Cat A")
    cat_b = repo.seed_category(code="B", name="Cat B")
    repo.seed_investment(
        exercise_id=eid,
        cost_center_id="205",
        category_id=cat_a,
        amount="100.00",
        priority="1",
        origin="national",
        required_date="2027-01-15",
    )
    repo.seed_investment(
        exercise_id=eid,
        cost_center_id="301",
        category_id=cat_b,
        amount="300.00",
        priority="3",
        origin="imported",
        required_date="2027-03-20",
    )
    repo.seed_plan(exercise_id=eid, cost_center_id="205", status="draft")
    repo.seed_plan(exercise_id=eid, cost_center_id="301", status="approved")

    by_unit = uc.get_grouping(_admin(), group_by="unit", exercise_id=eid)
    assert len(by_unit["items"]) == 2
    assert sum(i["investment_count"] for i in by_unit["items"]) == 2

    by_area = uc.get_grouping(_admin(), group_by="area", exercise_id=eid)
    assert {i["code"] for i in by_area["items"]} == {"PROD", "ENG"}

    by_cc = uc.get_grouping(_admin(), group_by="cost_center", exercise_id=eid)
    assert len(by_cc["items"]) == 2
    assert all("plan_status" in i for i in by_cc["items"])

    by_cat = uc.get_grouping(_admin(), group_by="category", exercise_id=eid)
    assert len(by_cat["items"]) == 2
    assert by_cat["items"][0]["percent_of_total"] is not None

    by_pri = uc.get_grouping(_admin(), group_by="priority", exercise_id=eid)
    assert {i["code"] for i in by_pri["items"]} == {"1", "3"}

    by_ori = uc.get_grouping(_admin(), group_by="origin", exercise_id=eid)
    assert {i["code"] for i in by_ori["items"]} == {"national", "imported"}

    by_month = uc.get_grouping(_admin(), group_by="month", exercise_id=eid)
    assert {i["code"] for i in by_month["items"]} == {"2027-01", "2027-03"}

    by_status = uc.get_grouping(_admin(), group_by="plan_status", exercise_id=eid)
    assert {i["code"] for i in by_status["items"]} == {"draft", "approved"}


def test_combined_filters_and_archived_ignored(uc, repo):
    eid = repo.seed_exercise()
    cat = repo.seed_category()
    repo.seed_investment(
        exercise_id=eid,
        cost_center_id="205",
        category_id=cat,
        amount="10.00",
        priority="2",
        origin="national",
        required_date="2027-05-01",
    )
    repo.seed_investment(
        exercise_id=eid,
        cost_center_id="205",
        category_id=cat,
        amount="999.00",
        status="archived",
    )
    repo.seed_investment(
        exercise_id=eid,
        cost_center_id="210",
        category_id=cat,
        amount="50.00",
        priority="4",
        origin="imported",
        required_date="2027-08-01",
    )
    result = uc.get_summary(
        _admin(),
        exercise_id=eid,
        unit_id="01",
        priority="2",
        origin="national",
        required_date_from="2027-01-01",
        required_date_to="2027-06-30",
        audit=False,
    )
    assert result["summary"]["investment_count"] == 1
    assert result["summary"]["total_estimated_amount"] == "10.00"


def test_plan_statuses_and_incomplete(uc, repo):
    eid = repo.seed_exercise()
    cat = repo.seed_category()
    repo.seed_investment(exercise_id=eid, cost_center_id="205", category_id=cat, amount="100.00")
    repo.seed_investment(
        exercise_id=eid, cost_center_id="210", category_id=cat, amount="200.00", incomplete=True
    )
    repo.seed_plan(exercise_id=eid, cost_center_id="205", status="changes_requested")
    repo.seed_plan(exercise_id=eid, cost_center_id="210", status="rejected")
    s = uc.get_summary(_admin(), exercise_id=eid, audit=False)["summary"]
    assert s["plans_changes_requested_count"] == 1
    assert s["plans_rejected_count"] == 1
    assert s["incomplete_investment_count"] == 1
    assert Decimal(s["in_review_amount"]) == Decimal("100.00")


def test_manager_scoped_and_idor(uc, repo):
    eid = repo.seed_exercise()
    cat = repo.seed_category()
    repo.seed_responsibility(user_sub="mgr", exercise_id=eid, cost_center_id="205")
    repo.seed_investment(exercise_id=eid, cost_center_id="205", category_id=cat, amount="100.00")
    repo.seed_investment(exercise_id=eid, cost_center_id="210", category_id=cat, amount="500.00")
    s = uc.get_summary(_manager(), exercise_id=eid, audit=False)["summary"]
    assert s["investment_count"] == 1
    assert s["total_estimated_amount"] == "100.00"

    with pytest.raises(CapexInvestmentCostCenterForbiddenError):
        uc.get_summary(_manager(), exercise_id=eid, cost_center_id="210", audit=False)


def test_auth_consolidation_and_export(uc, repo):
    eid = repo.seed_exercise()
    with pytest.raises((CapexConsolidationForbiddenError, BudgetUserNotAuthorizedError)):
        uc.get_summary(_no_perm(), exercise_id=eid, audit=False)

    with pytest.raises(CapexExportForbiddenError):
        uc.export_xlsx(_manager(), exercise_id=eid)

    # consolidator can export
    result = uc.export_xlsx(_consolidator(), exercise_id=eid)
    assert result["filename"].startswith("planejamento-capex-2027-")
    assert result["filename"].endswith(".xlsx")
    assert result["exported_count"] == 0
    assert repo.audits[-1]["action"] == "capex_consolidation.exported"


def test_currency_conflict(uc, repo):
    eid = repo.seed_exercise()
    cat = repo.seed_category()
    repo.seed_investment(
        exercise_id=eid, cost_center_id="205", category_id=cat, amount="10.00", currency="BRL"
    )
    repo.seed_investment(
        exercise_id=eid, cost_center_id="210", category_id=cat, amount="10.00", currency="USD"
    )
    with pytest.raises(CapexConsolidationCurrencyConflictError):
        uc.get_summary(_admin(), exercise_id=eid, audit=False)


def test_details_pagination(uc, repo):
    eid = repo.seed_exercise()
    cat = repo.seed_category()
    for i in range(3):
        repo.seed_investment(
            exercise_id=eid,
            cost_center_id="205",
            category_id=cat,
            amount=f"{(i + 1) * 100}.00",
            description=f"Item {i}",
        )
    page1 = uc.list_details(_admin(), exercise_id=eid, page=1, page_size=2)
    assert len(page1["items"]) == 2
    assert page1["pagination"]["total"] == 3
    assert page1["pagination"]["total_pages"] == 2
    assert "plan_status" in page1["items"][0]
    assert "is_complete" in page1["items"][0]


def test_excel_structure(uc, repo):
    eid = repo.seed_exercise()
    cat = repo.seed_category(code="FERR", name="Ferramentas")
    repo.seed_investment(
        exercise_id=eid,
        cost_center_id="205",
        category_id=cat,
        amount="1500.25",
        required_date="2027-04-10",
        description="Torno CNC",
    )
    repo.seed_plan(exercise_id=eid, cost_center_id="205", status="submitted", submitted_by="mgr")
    result = uc.export_xlsx(_consolidator(), exercise_id=eid, unit_id="01")
    raw = result["stream"].read()
    assert len(raw) > 1000
    wb = load_workbook(BytesIO(raw))
    assert wb.sheetnames == [
        "Resumo",
        "Investimentos",
        "Por Centro de Custo",
        "Por Categoria",
        "Por Mês",
    ]
    ws_inv = wb["Investimentos"]
    headers = [c.value for c in ws_inv[1]]
    assert "Valor previsto" in headers
    assert "Data Rcbto" in headers
    assert ws_inv.freeze_panes == "A2"
    assert ws_inv.auto_filter.ref
    # linha de dados
    amount_col = headers.index("Valor previsto") + 1
    date_col = headers.index("Data Rcbto") + 1
    assert isinstance(ws_inv.cell(2, amount_col).value, (int, float))
    date_val = ws_inv.cell(2, date_col).value
    assert date_val == date(2027, 4, 10) or (
        hasattr(date_val, "date") and date_val.date() == date(2027, 4, 10)
    )
    # resumo contém filtros
    resumo_values = [ws_inv.parent["Resumo"].cell(r, 2).value for r in range(1, 10)]
    assert any(isinstance(v, str) and "unit_id=01" in v for v in resumo_values)


def test_grouping_by_cost_center_keeps_same_code_per_branch_separate(uc, repo):
    """Mesmo código de CC em 01 e 02 não soma valores na consolidação."""
    eid = repo.seed_exercise()
    cat = repo.seed_category()
    repo.ccs[("02", "205")] = {
        **deepcopy(repo.ccs[("01", "205")]),
        "branch": "02",
        "unit_code": "02",
        "name": "TI ES",
    }
    for unit_id, amount, desc in (
        ("01", "8000.00", "Jaraguá"),
        ("02", "122600.00", "Rio Bananal"),
    ):
        iid = str(uuid4())
        repo.investments[iid] = {
            "id": iid,
            "exercise_id": eid,
            "unit_id": unit_id,
            "area_id": "PROD",
            "cost_center_id": "205",
            "category_id": cat,
            "description": desc,
            "estimated_amount": amount,
            "currency": "BRL",
            "required_date": "2027-06-01",
            "priority": "2",
            "origin": "national",
            "probable_supplier_name": "Fornecedor X",
            "status": "draft",
            "review_status": "pending",
            "version": 1,
            "updated_at": "2026-08-01T10:00:00Z",
            "created_at": "2026-07-01T10:00:00Z",
        }

    by_cc = uc.get_grouping(_admin(), group_by="cost_center", exercise_id=eid)
    assert len(by_cc["items"]) == 2
    by_pair = {
        (i.get("unit_id"), i.get("cost_center_id")): Decimal(str(i["total_amount"]))
        for i in by_cc["items"]
    }
    assert by_pair[("01", "205")] == Decimal("8000.00")
    assert by_pair[("02", "205")] == Decimal("122600.00")
