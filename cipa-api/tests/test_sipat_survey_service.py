from __future__ import annotations

from io import BytesIO
from types import SimpleNamespace
from unittest.mock import MagicMock

import pytest

from cipa_app.application.use_cases.sipat_survey_service import SipatSurveyService


def _user(*perms: str):
    return SimpleNamespace(id="user-1", permissions=list(perms))


def test_normalize_questions_requires_options_for_choice():
    svc = SipatSurveyService(repository=MagicMock(), qr_service=MagicMock())
    with pytest.raises(ValueError, match="ao menos 2 opções"):
        svc._normalize_questions(
            [{"question_type": "single_choice", "label": "X", "options": ["a"]}]
        )


def test_create_and_publish_with_template(monkeypatch):
    repo = MagicMock()
    qr = MagicMock()
    qr.generate.return_value = "sipat-tok.png"
    survey = {
        "id": "s1",
        "unit_code": "01",
        "title": "SIPAT 2026",
        "description": None,
        "status": "draft",
        "public_token": None,
        "qr_filename": None,
        "opens_at": None,
        "closes_at": None,
        "response_count": 0,
        "created_at": None,
        "updated_at": None,
    }
    published = {
        **survey,
        "status": "published",
        "public_token": "tok",
        "qr_filename": "sipat-tok.png",
    }

    def create_survey(payload):
        return {**survey, "title": payload["title"], "description": payload.get("description")}

    repo.create_survey.side_effect = create_survey
    repo.replace_questions.return_value = [
        {
            "id": "q1",
            "position": 0,
            "question_type": "yes_no",
            "label": "Ok?",
            "help_text": None,
            "is_required": True,
            "options": ["Sim", "Não"],
        }
    ]
    repo.get_survey.return_value = survey
    repo.list_questions.return_value = repo.replace_questions.return_value
    repo.update_survey.return_value = published

    svc = SipatSurveyService(repository=repo, qr_service=qr)
    user = _user("cipa.sipat.manage", "cipa.unit.filial-01")

    created = svc.create_survey(
        user,
        {
            "unit_code": "01",
            "title": "SIPAT 2026",
            "template_id": "seguranca_rapida",
        },
    )
    assert created["survey"]["title"] == "SIPAT 2026"
    assert len(created["questions"]) >= 1

    repo.get_survey.return_value = survey
    result = svc.publish(user, "s1")
    assert result["survey"]["status"] == "published"
    qr.generate.assert_called_once()


def test_submit_public_rejects_closed():
    repo = MagicMock()
    repo.get_by_token.return_value = {
        "id": "s1",
        "status": "closed",
        "opens_at": None,
        "closes_at": None,
    }
    svc = SipatSurveyService(repository=repo, qr_service=MagicMock())
    with pytest.raises(ValueError, match="não está aberta"):
        svc.submit_public("tok", {"answers": []})


def test_submit_public_ok():
    repo = MagicMock()
    repo.get_by_token.return_value = {
        "id": "s1",
        "status": "published",
        "opens_at": None,
        "closes_at": None,
    }
    repo.list_questions.return_value = [
        {
            "id": "q1",
            "question_type": "yes_no",
            "label": "Ok?",
            "is_required": True,
            "options": ["Sim", "Não"],
        }
    ]
    repo.create_response_with_answers.return_value = {"id": "r1"}
    svc = SipatSurveyService(repository=repo, qr_service=MagicMock())
    result = svc.submit_public(
        "tok",
        {"answers": [{"question_id": "q1", "value": "Sim"}]},
    )
    assert result["ok"] is True
    assert result["response_id"] == "r1"


def test_list_surveys_requires_unit_rbac():
    repo = MagicMock()
    svc = SipatSurveyService(repository=repo, qr_service=MagicMock())
    with pytest.raises(PermissionError):
        svc.list_surveys(_user("cipa.sipat.view"), unit_code="01")


def test_clone_survey_copies_questions_as_draft():
    repo = MagicMock()
    source = {
        "id": "s1",
        "unit_code": "01",
        "title": "SIPAT 2025",
        "description": "Desc original",
        "status": "published",
        "public_token": "tok",
        "qr_filename": "sipat-tok.png",
        "opens_at": None,
        "closes_at": None,
        "response_count": 12,
        "created_at": None,
        "updated_at": None,
    }
    clone = {
        "id": "s2",
        "unit_code": "01",
        "title": "Cópia de SIPAT 2025",
        "description": "Desc original",
        "status": "draft",
        "public_token": None,
        "qr_filename": None,
        "opens_at": None,
        "closes_at": None,
        "response_count": 0,
        "created_at": None,
        "updated_at": None,
    }
    repo.get_survey.return_value = source
    repo.list_questions.return_value = [
        {
            "id": "q1",
            "position": 0,
            "question_type": "likert_5",
            "label": "Satisfação?",
            "help_text": "1–5",
            "is_required": True,
            "options": ["1", "2", "3", "4", "5"],
        },
        {
            "id": "q2",
            "position": 1,
            "question_type": "text_long",
            "label": "Comentário",
            "help_text": None,
            "is_required": False,
            "options": None,
        },
    ]
    repo.create_survey.return_value = clone
    repo.replace_questions.return_value = [
        {**q, "id": f"new-{q['id']}"} for q in repo.list_questions.return_value
    ]

    svc = SipatSurveyService(repository=repo, qr_service=MagicMock())
    user = _user("cipa.sipat.manage", "cipa.unit.filial-01")
    result = svc.clone_survey(user, "s1")

    assert result["survey"]["id"] == "s2"
    assert result["survey"]["status"] == "draft"
    assert result["survey"]["title"] == "Cópia de SIPAT 2025"
    assert result["survey"]["public_token"] is None
    assert len(result["questions"]) == 2
    create_payload = repo.create_survey.call_args.args[0]
    assert create_payload["unit_code"] == "01"
    assert create_payload["title"] == "Cópia de SIPAT 2025"
    assert create_payload["description"] == "Desc original"
    normalized = repo.replace_questions.call_args.args[1]
    assert normalized[0]["label"] == "Satisfação?"
    assert normalized[0]["question_type"] == "likert_5"
    assert normalized[1]["is_required"] is False


def test_public_prefix_is_public():
    from cipa_app.middleware.auth_middleware import _is_public

    assert _is_public("/public/sipat/abc") is True
    assert _is_public("/sipat/surveys") is False


def test_build_sipat_results_xlsx_has_summary_and_responses():
    from openpyxl import load_workbook

    from cipa_app.application.services.sipat_excel_export_service import (
        build_sipat_results_xlsx,
        group_response_answers,
    )

    survey = {
        "id": "s1",
        "title": "SIPAT 2026 ES",
        "description": "Clima",
        "unit_code": "01",
        "status": "published",
        "response_count": 2,
    }
    questions = [
        {
            "id": "q1",
            "label": "Satisfação?",
            "question_type": "likert_5",
        },
        {
            "id": "q2",
            "label": "Comentário",
            "question_type": "text_long",
        },
    ]
    summary_questions = [
        {
            "question_id": "q1",
            "label": "Satisfação?",
            "question_type": "likert_5",
            "answer_count": 2,
            "counts": {"5": 1, "4": 1},
        },
        {
            "question_id": "q2",
            "label": "Comentário",
            "question_type": "text_long",
            "answer_count": 1,
            "sample_texts": ["Ótimo"],
        },
    ]
    responses = group_response_answers(
        [
            {
                "response_id": "r1",
                "created_at": None,
                "question_id": "q1",
                "value_text": "5",
                "value_json": None,
            },
            {
                "response_id": "r1",
                "created_at": None,
                "question_id": "q2",
                "value_text": "Ótimo",
                "value_json": None,
            },
            {
                "response_id": "r2",
                "created_at": None,
                "question_id": "q1",
                "value_text": "4",
                "value_json": None,
            },
        ]
    )
    raw, filename = build_sipat_results_xlsx(
        survey=survey,
        questions=questions,
        summary_questions=summary_questions,
        responses=responses,
    )
    assert filename.endswith(".xlsx")
    assert "sipat-" in filename
    wb = load_workbook(BytesIO(raw))
    assert wb.sheetnames == ["Resumo", "Respostas"]
    assert wb["Respostas"]["A2"].value == 1
    assert wb["Respostas"]["B2"].value == "5"
    assert wb["Respostas"]["C2"].value == "Ótimo"
    assert wb["Respostas"]["A3"].value == 2
    assert wb["Respostas"]["B3"].value == "4"


def test_export_excel_requires_view_permission():
    repo = MagicMock()
    repo.get_survey.return_value = {
        "id": "s1",
        "unit_code": "01",
        "title": "X",
        "description": None,
        "status": "published",
        "public_token": "tok",
        "qr_filename": None,
        "opens_at": None,
        "closes_at": None,
        "response_count": 0,
        "created_at": None,
        "updated_at": None,
    }
    svc = SipatSurveyService(repository=repo, qr_service=MagicMock())
    with pytest.raises(PermissionError):
        svc.export_excel(_user("cipa.sipat.view"), "s1")
