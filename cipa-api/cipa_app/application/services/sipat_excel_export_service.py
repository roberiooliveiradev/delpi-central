from __future__ import annotations

import re
from io import BytesIO
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter

QUESTION_TYPE_LABELS = {
    "single_choice": "Escolha única",
    "multi_choice": "Múltipla escolha",
    "likert_5": "Escala 1–5",
    "yes_no": "Sim/Não",
    "text_short": "Texto curto",
    "text_long": "Texto longo",
}

CHOICE_TYPES = frozenset({"single_choice", "multi_choice", "likert_5", "yes_no"})


def _slugify(value: str, *, max_len: int = 40) -> str:
    text = re.sub(r"[^\w\s-]", "", (value or "").strip(), flags=re.UNICODE)
    text = re.sub(r"[-\s]+", "-", text).strip("-").lower()
    return (text or "sipat")[:max_len]


def _answer_display(value_text: Any, value_json: Any) -> str:
    if value_json is not None:
        if isinstance(value_json, list):
            return "; ".join(str(item) for item in value_json if str(item).strip())
        return str(value_json)
    if value_text is None:
        return ""
    return str(value_text).strip()


def _autosize(ws, *, max_width: int = 48) -> None:
    for column_cells in ws.columns:
        letter = get_column_letter(column_cells[0].column)
        length = 0
        for cell in column_cells:
            value = "" if cell.value is None else str(cell.value)
            length = max(length, min(len(value), max_width))
        ws.column_dimensions[letter].width = max(12, length + 2)


def build_sipat_results_xlsx(
    *,
    survey: dict[str, Any],
    questions: list[dict[str, Any]],
    summary_questions: list[dict[str, Any]],
    responses: list[dict[str, Any]],
) -> tuple[bytes, str]:
    """Gera Excel anônimo: Resumo + Respostas (uma linha por participação)."""
    wb = Workbook()
    header_font = Font(bold=True)
    wrap = Alignment(wrap_text=True, vertical="top")

    meta = wb.active
    meta.title = "Resumo"
    meta.append(["Campo", "Valor"])
    for cell in meta[1]:
        cell.font = header_font
    meta.append(["Título", survey.get("title") or ""])
    meta.append(["Descrição", survey.get("description") or ""])
    meta.append(["Unidade", survey.get("unit_code") or ""])
    meta.append(["Status", survey.get("status") or ""])
    meta.append(["Total de respostas", int(survey.get("response_count") or 0)])
    meta.append([])
    meta.append(["Pergunta", "Tipo", "Opção / texto", "Quantidade", "%"])
    for cell in meta[meta.max_row]:
        cell.font = header_font

    summary_by_id = {str(item.get("question_id")): item for item in summary_questions}
    for index, question in enumerate(questions, start=1):
        qid = str(question["id"])
        label = f"{index}. {question.get('label') or ''}"
        qtype = str(question.get("question_type") or "")
        type_label = QUESTION_TYPE_LABELS.get(qtype, qtype)
        entry = summary_by_id.get(qid) or {}
        answer_count = int(entry.get("answer_count") or 0)
        if qtype in CHOICE_TYPES:
            counts = entry.get("counts") or {}
            if not counts:
                meta.append([label, type_label, "(sem respostas)", 0, 0])
                continue
            ordered = sorted(
                counts.items(),
                key=lambda item: (
                    int(item[0]) if qtype == "likert_5" and str(item[0]).isdigit() else 0,
                    -int(item[1]),
                    str(item[0]),
                ),
            )
            for option, count in ordered:
                share = round((count / answer_count) * 100, 1) if answer_count else 0
                meta.append([label, type_label, str(option), int(count), share])
                label = ""
                type_label = ""
        else:
            texts = entry.get("sample_texts") or []
            if not texts:
                meta.append([label, type_label, "(sem respostas)", 0, 0])
                continue
            for text in texts:
                meta.append([label, type_label, str(text), 1, ""])
                label = ""
                type_label = ""

    for row in meta.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = wrap
    _autosize(meta)

    answers = wb.create_sheet("Respostas")
    headers = ["#"]
    for index, question in enumerate(questions, start=1):
        headers.append(f"{index}. {question.get('label') or ''}")
    answers.append(headers)
    for cell in answers[1]:
        cell.font = header_font
        cell.alignment = wrap

    # responses: [{id, created_at, answers: {question_id: display}}]
    for row_index, response in enumerate(responses, start=1):
        by_qid = response.get("answers") or {}
        row = [row_index]
        for question in questions:
            row.append(by_qid.get(str(question["id"]), ""))
        answers.append(row)

    for row in answers.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = wrap
    _autosize(answers, max_width=56)

    buffer = BytesIO()
    wb.save(buffer)
    slug = _slugify(str(survey.get("title") or "sipat"))
    filename = f"sipat-{slug}-resultados.xlsx"
    return buffer.getvalue(), filename


def group_response_answers(
    flat_rows: list[dict[str, Any]],
) -> list[dict[str, Any]]:
    """Agrupa linhas planas (response_id, question_id, value_*) em respostas."""
    ordered: list[dict[str, Any]] = []
    index: dict[str, dict[str, Any]] = {}
    for row in flat_rows:
        rid = str(row["response_id"])
        bucket = index.get(rid)
        if bucket is None:
            bucket = {
                "id": rid,
                "created_at": row.get("created_at"),
                "answers": {},
            }
            index[rid] = bucket
            ordered.append(bucket)
        qid = row.get("question_id")
        if qid is None:
            continue
        bucket["answers"][str(qid)] = _answer_display(
            row.get("value_text"), row.get("value_json")
        )
    return ordered
